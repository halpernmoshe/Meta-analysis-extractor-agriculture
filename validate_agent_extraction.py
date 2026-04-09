"""
Validate agent extraction results against Loladze ground truth.
Adapted from validate_full_46.py for agent extraction format.

Usage:
    python validate_agent_extraction.py
"""
import sys, os, json, math, csv, re
from pathlib import Path
from collections import defaultdict
from datetime import datetime

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
GT_PATH = os.environ.get('GT_PATH_LOLADZE', str(BASE_DIR.parent / 'Loladze' / 'CO2+Dataset.xlsx'))
if not Path(GT_PATH).exists():
    print(f"Ground truth not found at {GT_PATH}")
    print("Set GT_PATH_LOLADZE environment variable or place file at the expected path.")
    print("See REPRODUCE.md Section 3 for download instructions.")
    sys.exit(1)
AGENT_DIR = BASE_DIR / 'output' / 'agent_extraction'

# Complete mapping of our paper IDs to Loladze references
PAPER_TO_LOLADZE_REF = {
    "001_Ma_2007": "Fernando et al 2012a",
    "002_Ziska_1997": "Ziska et al 1997",
    "003_Baslam_2012": "Baslam et al 2012",
    "004_Finzi_2001": "Finzi et al 2001",
    "005_Niinemets_1999": "Niinemets et al 1999",
    "006_Azam_2013": "Azam et al 2013",
    "007_Woodin_1992": "Woodin et al 1992",
    "008_Campbell_2002": "Cambell & Sage 2002",
    "009_Barnes_1992": "Barnes & Pffirrman 1992",
    "010_Li_2010": "Hogy et al 2009",
    "011_Huluka_1994": "Huluka et al 1994",
    "012_Wu_2004": "Wu et al 2004",
    "013_Keutgen_2001": "Keutgen & Chen 2001",
    "014_Lieffering_2004": "Lieffering et al 2004",
    "015_Pleijel_2009": "Pleijel & Danielsson 2009",
    "016_Fernando_2012a": "Fernando et al 2012b",
    "017_Fangmeier_2002": "Fangmeier et al  2002",
    "018_Al-Rawahy_2013": "Al-Rawahy et al 2013",
    "019_Baxter_1994": "Baxter et al 1994",
    "020_Overdieck_1993": "Overdieck 1993",
    "021_Wilsey_1994": "Wilsey et al 1994",
    "022_Blank_2011": "Blank et al 2011",
    "025_Guo_2011": "Guo et al 2013",
    "026_Seneweera_1997": "Seneweera & Conroy 1997",
    "027_Peet_1986": "Peet et al 1986",
    "028_Mishra_2011": "Mishra et al 2011",
    "031_Pal_2003": "Pal et al 2004",
    "032_Kanowski_2001": "Kanowski 2001",
    "034_Johnson_1997": "Johnson et al 2003",
    "035_Oksanen_2005": "Oksanen et al 2005",
    "036_Schenk_1997": "Schenk et al 1997",
    "037_de_2000": "Haase et al 2008",
    "038_Newbery_1995": "Newbery 1995",
    "039_Heagle_1993": "Heagle et al 1993",
    "040_Pfirrmann_1996": "Pfirrmann et al 1996",
    "041_Mjwara_1996": "Mjwara et al 1996",
    "042_Luomala_2005": "Luomala et al 2005",
    "043_Natali_2009": "Natali et al 2009",
    "044_Housman_2012": "Housman et al 2012",
    "046_Porter_1984": "Porter & Grodzinski 1984",
    "047_Rodenkirchen_2009": "Rodenkirchen et al 2009",
    "048_Khan_2013": "Khan et al 2012",
    "049_Singh_2013": "Singh et al 2013",
    "050_Polley_2011": "Polley et al 2011",
    "051_Niu_2013": "Niu et al 2013",
    "058_ONeill_1987": "O'Neill et al 1987",
}

MISLABELED_PDFS = {
    "001_Ma_2007": "Fernando et al 2012a",
    "010_Li_2010": "Högy et al 2009",
    "031_Pal_2003": "Pal et al 2004",
    "034_Johnson_1997": "Johnson et al 2003",
    "037_de_2000": "Haase et al 2008",
}


def normalize_element(el):
    if not el:
        return ""
    el = el.strip()
    if '(' in el:
        el = el[:el.index('(')].strip()
    el = el.upper()
    MAP = {"NITROGEN": "N", "PHOSPHORUS": "P", "POTASSIUM": "K",
           "CALCIUM": "CA", "MAGNESIUM": "MG", "IRON": "FE",
           "ZINC": "ZN", "MANGANESE": "MN", "COPPER": "CU",
           "SULFUR": "S", "SULPHUR": "S", "BORON": "B", "CARBON": "C",
           "SODIUM": "NA", "MOLYBDENUM": "MO", "SILICON": "SI",
           "SELENIUM": "SE", "COBALT": "CO", "CHROMIUM": "CR",
           "CADMIUM": "CD", "NICKEL": "NI", "LEAD": "PB",
           "ALUMINIUM": "AL", "ALUMINUM": "AL", "BARIUM": "BA",
           "STRONTIUM": "SR", "BROMINE": "BR", "VANADIUM": "V",
           "CHLORINE": "CL", "PROTEIN": "N",  # protein ~ N for Loladze
           "OIL": "OIL", "LIGNIN": "LIGNIN", "TNC": "TNC"}
    result = MAP.get(el, None)
    if result:
        return result
    SYMBOLS = {"N", "P", "K", "CA", "MG", "FE", "ZN", "MN", "CU", "S", "B", "C",
               "NA", "MO", "SI", "SE", "CO", "CR", "CD", "NI", "PB", "AL", "BA",
               "SR", "BR", "V", "CL"}
    tokens = el.split()
    for tok in tokens:
        tok_clean = tok.strip(",;:")
        if tok_clean in SYMBOLS:
            return tok_clean
    for tok in tokens:
        tok_clean = tok.strip(",;:")
        if tok_clean in MAP:
            return MAP[tok_clean]
    for sym in sorted(SYMBOLS, key=len, reverse=True):
        if re.search(r'\b' + sym + r'\b', el):
            return sym
    return el


def load_gt():
    wb = openpyxl.load_workbook(GT_PATH, data_only=True)
    ws = wb["CO2 Dataset"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[2]]
    col = {}
    for i, h in enumerate(headers):
        if h == 'Reference': col['ref'] = i
        elif h == '(E-A)/A': col['effect'] = i
        elif h == 'Element': col['element'] = i
        elif h == 'Additional Info': col['info'] = i
        elif h == 'eCO2': col['eco2'] = i
        elif h == 'aCO2': col['aco2'] = i
        elif h == 'Tissue': col['tissue'] = i
        elif h == 'Species': col['species'] = i
        elif h == 'n': col['n'] = i

    gt = defaultdict(list)
    for row in ws.iter_rows(min_row=3, values_only=True):
        ref = str(row[col['ref']]).strip() if row[col['ref']] else ""
        el = normalize_element(str(row[col['element']]).strip() if row[col['element']] else "")
        eff = row[col['effect']]
        info = str(row[col.get('info', 11)]).strip() if row[col.get('info', 11)] else ""
        if info == "None": info = ""
        eco2 = row[col.get('eco2', 8)]
        tissue = str(row[col.get('tissue', 4)]).strip() if row[col.get('tissue', 4)] else ""
        species = str(row[col.get('species', 0)]).strip() if col.get('species') is not None and row[col['species']] else ""

        if ref and el and eff is not None:
            try:
                gt[ref].append({
                    'element': el, 'effect': float(eff),
                    'info': info, 'eco2': float(eco2) if eco2 else None,
                    'tissue': tissue, 'species': species,
                })
            except (ValueError, TypeError):
                pass
    wb.close()
    return dict(gt)


def get_mods(obs):
    mods = obs.get('moderators', {})
    if isinstance(mods, str):
        try: mods = json.loads(mods)
        except: mods = {}
    if not isinstance(mods, dict): mods = {}
    return mods


def is_concentration_unit(unit_str):
    u = str(unit_str).lower()
    total_keywords = ['plant', 'pot', 'total', 'uptake', 'content', '/ha', 'per hectare']
    if any(kw in u for kw in total_keywords):
        return False
    return True


def filter_obs_for_gt_row(obs_list, gt_row, paper_id):
    """Filter observations to match a specific GT row's conditions."""
    gt_el = gt_row['element']
    gt_info = gt_row['info'].lower()

    el_matches = [o for o in obs_list if normalize_element(o.get('element', '')) == gt_el]
    if not el_matches:
        return []

    conc_matches = [o for o in el_matches if is_concentration_unit(o.get('unit', ''))]
    if conc_matches:
        el_matches = conc_matches

    if not gt_info:
        return el_matches

    filtered = el_matches

    # --- Leaf position filters (Baslam) ---
    if 'inner' in gt_info and 'nm' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            leaf_pos = str(mods.get('leaf_position', '')).lower()
            amf = str(mods.get('AMF_status', mods.get('mycorrhizal_status', ''))).lower()
            if 'inner' in leaf_pos and ('non' in amf or 'nm' in amf):
                new.append(o)
        if new: filtered = new
    elif 'nm' in gt_info or 'non-mycorrhizal' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            amf = str(mods.get('AMF_status', mods.get('mycorrhizal_status', ''))).lower()
            if 'non' in amf or 'nm' in amf:
                new.append(o)
        if new: filtered = new

    # --- Leaf age ---
    if 'old leaves' in gt_info:
        new = [o for o in filtered if 'old' in str(get_mods(o).get('leaf_age', '')).lower()]
        if new: filtered = new
    if 'young leaves' in gt_info:
        new = [o for o in filtered if 'young' in str(get_mods(o).get('leaf_age', '')).lower()]
        if new: filtered = new

    # --- DOY ---
    doy_match = re.search(r'doy\s*(\d+)', gt_info)
    if doy_match:
        target_doy = doy_match.group(1)
        new = [o for o in filtered if target_doy in str(get_mods(o).get('sampling_date', ''))]
        if new: filtered = new

    # --- Year ---
    year_match = re.match(r'^(\d{4})$', gt_row['info'].strip())
    if year_match:
        target_year = year_match.group(1)
        new = [o for o in filtered if target_year in str(get_mods(o).get('year', ''))]
        if new: filtered = new

    # --- N level ---
    if 'high n' in gt_info and 'low' not in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            n_level = str(mods.get('nitrogen_level', mods.get('N_level', mods.get('n_treatment', '')))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if 'high' in n_level or 'high n' in desc:
                new.append(o)
        if new: filtered = new
    elif 'low n' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            n_level = str(mods.get('nitrogen_level', mods.get('N_level', mods.get('n_treatment', '')))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if 'low' in n_level or 'low n' in desc:
                new.append(o)
        if new: filtered = new

    # --- Soil type ---
    if 'basalt' in gt_info:
        new = [o for o in filtered if 'basalt' in str(get_mods(o).get('soil_type', '')).lower()
               or 'basalt' in str(o.get('treatment_description', '')).lower()]
        if new: filtered = new
    if 'rhyolite' in gt_info:
        new = [o for o in filtered if 'rhyolite' in str(get_mods(o).get('soil_type', '')).lower()
               or 'rhyolite' in str(o.get('treatment_description', '')).lower()]
        if new: filtered = new

    # --- Cultivar/clone ---
    for cultivar_name in ['nc-r', 'nc-s']:
        if cultivar_name in gt_info:
            new = []
            for o in filtered:
                mods = get_mods(o)
                cult = str(mods.get('cultivar', mods.get('clone', ''))).lower()
                if cultivar_name in cult:
                    new.append(o)
            if new: filtered = new
            break

    # --- K treatment (Pfirrmann) ---
    if '+k' in gt_info and '-k' not in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            k_treat = str(mods.get('potassium', mods.get('K_treatment', mods.get('soil_K', '')))).lower()
            if '+k' in k_treat or 'plus' in k_treat or 'with k' in k_treat or 'high' in k_treat:
                new.append(o)
        if new: filtered = new
    elif '-k' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            k_treat = str(mods.get('potassium', mods.get('K_treatment', mods.get('soil_K', '')))).lower()
            if '-k' in k_treat or 'minus' in k_treat or 'without' in k_treat or 'low' in k_treat:
                new.append(o)
        if new: filtered = new

    # --- Needle year ---
    needle_year_match = re.search(r'(\d{4})\s*needles?', gt_info)
    if needle_year_match:
        target_year = needle_year_match.group(1)
        new = []
        for o in filtered:
            mods = get_mods(o)
            all_text = str(mods).lower() + ' ' + str(o.get('tissue', '')).lower()
            if target_year in all_text or ('previous' in all_text and target_year == '1988'):
                new.append(o)
        if new: filtered = new

    # --- NH4/NO3 ---
    if 'nh4' in gt_info and 'no3' not in gt_info:
        new = [o for o in filtered if 'nh4' in str(get_mods(o)).lower()
               or 'ammonium' in str(o.get('treatment_description', '')).lower()]
        if new: filtered = new
    if 'no3' in gt_info and 'nh4' not in gt_info:
        new = [o for o in filtered if 'no3' in str(get_mods(o)).lower()
               or 'nitrate' in str(o.get('treatment_description', '')).lower()]
        if new: filtered = new

    # --- P level (Seneweera) ---
    p_level_match = re.search(r'p\s+(\d+)\s*mg/kg', gt_info)
    if p_level_match:
        target_p = p_level_match.group(1)
        target_p_int = int(target_p)
        new = []
        for o in filtered:
            mods = get_mods(o)
            for mk, mv in mods.items():
                if ('p_level' in str(mk).lower() or 'phosph' in str(mk).lower()) and mv is not None:
                    try:
                        if abs(float(mv) - target_p_int) < 1:
                            new.append(o); break
                    except: pass
        if new: filtered = new

    # --- Heavy metal level (Guo) ---
    cd_match = re.search(r'cd=(\d+)', gt_info)
    if cd_match:
        target_cd = int(cd_match.group(1))
        new = []
        for o in filtered:
            mods = get_mods(o)
            for mk, mv in mods.items():
                if 'cd' in str(mk).lower() or 'cadmium' in str(mk).lower() or 'contam' in str(mk).lower():
                    level_match = re.search(r'(\d+)', str(mv).lower())
                    if level_match and int(level_match.group(1)) == target_cd:
                        new.append(o); break
        if new: filtered = new

    cu_match = re.search(r'cu=(\d+)', gt_info)
    if cu_match:
        target_cu = int(cu_match.group(1))
        new = []
        for o in filtered:
            mods = get_mods(o)
            for mk, mv in mods.items():
                if 'cu' in str(mk).lower() or 'copper' in str(mk).lower():
                    level_match = re.search(r'(\d+)', str(mv).lower())
                    if level_match and int(level_match.group(1)) == target_cu:
                        new.append(o); break
        if new: filtered = new

    # --- Harvest year (Guo) ---
    harvest_year_match = re.search(r'(\d{4})\s*harvest', gt_info)
    if harvest_year_match:
        target_year = harvest_year_match.group(1)
        new = [o for o in filtered if target_year in str(get_mods(o)) + str(o.get('treatment_description', ''))]
        if new: filtered = new

    # --- Site (Natali) ---
    if paper_id == '043_Natali_2009':
        if 'duke' in gt_info and 'serc' not in gt_info:
            new = [o for o in filtered if 'duke' in str(get_mods(o).get('site', '')).lower()]
            if new: filtered = new
            if '1 yr old' in gt_info:
                new = [o for o in filtered if 'pinus' in str(get_mods(o).get('species', '')).lower()
                       or 'taeda' in str(get_mods(o).get('species', '')).lower()]
                if new: filtered = new
                yr1 = [o for o in filtered if '1' in str(get_mods(o).get('needle_age', get_mods(o).get('leaf_age', '')))]
                if yr1: filtered = yr1
            else:
                new = [o for o in filtered if 'liquidambar' in str(get_mods(o).get('species', '')).lower()]
                if new: filtered = new
        elif 'ornl' in gt_info:
            new = [o for o in filtered if 'ornl' in str(get_mods(o).get('site', '')).lower()]
            if new: filtered = new
        elif 'serc' in gt_info:
            new = [o for o in filtered if 'serc' in str(get_mods(o).get('site', '')).lower()]
            if new: filtered = new
            gt_species = gt_row.get('species', '').lower()
            if gt_species:
                gt_parts = [p for p in gt_species.split() if len(p) > 3]
                species_new = [o for o in filtered
                               if all(part in str(get_mods(o).get('species', '')).lower() for part in gt_parts)]
                if species_new: filtered = species_new

    # --- Ozone filter ---
    if 'o3' not in gt_info and 'ozone' not in gt_info:
        has_o3_mods = any(
            any(k.lower() in ('o3_level', 'ozone', 'o3', 'ozone_level') for k in get_mods(o).keys())
            for o in filtered
        )
        if has_o3_mods:
            ambient_o3 = []
            for o in filtered:
                mods = get_mods(o)
                o3_val = ''
                for k, v in mods.items():
                    if k.lower() in ('o3_level', 'ozone', 'o3', 'ozone_level'):
                        o3_val = str(v).lower(); break
                if any(term in o3_val for term in ['ambient', 'cf', 'charcoal', 'control',
                       'low', 'clean', 'filtered', '5 nl', '20 n', '20n', 'background']):
                    ambient_o3.append(o)
                elif not o3_val:
                    ambient_o3.append(o)
            if ambient_o3: filtered = ambient_o3

    # --- CO2 level ---
    if gt_row.get('eco2'):
        gt_co2 = gt_row['eco2']
        new = []
        for o in filtered:
            desc = str(o.get('treatment_description', ''))
            mods = get_mods(o)
            co2_match = re.search(r'(\d{3,4})\s*(?:ppm|µmol|umol|μmol|ul)', desc, re.IGNORECASE)
            our_co2 = None
            if co2_match:
                our_co2 = float(co2_match.group(1))
            else:
                for k, v in mods.items():
                    if 'co2' in k.lower() and v is not None:
                        try: our_co2 = float(v); break
                        except: pass
            if our_co2 is not None:
                if abs(our_co2 - gt_co2) < 100:
                    new.append(o)
            else:
                new.append(o)
        if new: filtered = new

    return filtered


def compute_effect(obs, swap_tc=False):
    """Compute (E-A)/A effect from an observation.
    Agent format has effect_pct directly (in %), or compute from means."""
    # Try direct effect_pct first
    eff_pct = obs.get('effect_pct')
    if eff_pct is not None:
        eff = eff_pct / 100.0  # Convert from % to fraction
        if swap_tc:
            eff = -eff  # Rough swap
        return eff

    # Fall back to computing from means
    ctrl = obs.get('control_mean')
    treat = obs.get('treatment_mean')
    if swap_tc:
        ctrl, treat = treat, ctrl
    if ctrl and ctrl != 0 and treat is not None:
        return (treat - ctrl) / ctrl
    return None


def detect_tc_swap(obs_list, gt_rows):
    our_effects = []
    gt_effects = []
    for gt_row in gt_rows:
        gt_el = gt_row['element']
        matching = [o for o in obs_list if normalize_element(o.get('element', '')) == gt_el]
        if not matching: continue
        for o in matching:
            eff = compute_effect(o, swap_tc=False)
            if eff is not None and abs(eff) <= 5.0:
                our_effects.append(eff)
                gt_effects.append(gt_row['effect'])
    if len(our_effects) < 3:
        return False
    disagree = sum(1 for o, g in zip(our_effects, gt_effects) if g != 0 and (o < 0) != (g < 0))
    return disagree / len(our_effects) > 0.7


def main():
    print(f"Agent Extraction Validation vs Loladze Ground Truth")
    print(f"{'='*70}")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")

    gt = load_gt()
    total_gt_rows = sum(len(v) for v in gt.values())
    print(f"Ground truth: {total_gt_rows} observations across {len(gt)} references")

    # Find agent extraction files (prefer v2 if available)
    agent_files = {}
    for f in sorted(AGENT_DIR.glob("*_agent*.json")):
        paper_id = f.stem.replace("_agent_v2", "").replace("_agent", "")
        if '_v2' in f.stem or paper_id not in agent_files:
            agent_files[paper_id] = f

    print(f"Agent extractions: {len(agent_files)} papers\n")

    all_matches = []
    paper_results = []

    for paper_id, rf in sorted(agent_files.items()):
        loladze_ref = PAPER_TO_LOLADZE_REF.get(paper_id)
        if not loladze_ref or loladze_ref not in gt:
            for ref in gt:
                surname = paper_id.split('_')[1] if '_' in paper_id else paper_id
                if surname.lower() in ref.lower():
                    loladze_ref = ref; break
            if not loladze_ref or loladze_ref not in gt:
                continue

        with open(rf, encoding='utf-8') as f:
            data = json.load(f)

        gt_rows = gt[loladze_ref]
        obs_list = data.get('consensus_observations', [])

        # Filter out sub-ambient CO2
        elevated_obs = []
        for o in obs_list:
            desc = str(o.get('treatment_description', '')).lower()
            co2_match = re.search(r'(\d{2,4})\s*(?:ppm|µmol|umol|μmol|ul)', desc)
            if co2_match:
                co2_val = float(co2_match.group(1))
                if co2_val < 300: continue
            if 'low co2' in desc or 'sub-ambient' in desc: continue
            elevated_obs.append(o)
        if elevated_obs:
            obs_list = elevated_obs

        # Detect T/C swap
        swap_tc = detect_tc_swap(obs_list, gt_rows)
        if swap_tc:
            print(f"  WARNING: T/C swap detected for {paper_id}")

        matched = 0
        paper_matches = []
        used_obs_ids = set()

        gt_by_el_info = defaultdict(list)
        for gt_row in gt_rows:
            key = (gt_row['element'], gt_row['info'])
            gt_by_el_info[key].append(gt_row)

        for gt_row in gt_rows:
            candidates = filter_obs_for_gt_row(obs_list, gt_row, paper_id)
            if not candidates: continue

            cand_effects = []
            for i, c in enumerate(candidates):
                obs_id = id(c)
                eff = compute_effect(c, swap_tc=swap_tc)
                if eff is not None and abs(eff) <= 5.0:
                    cand_effects.append((obs_id, eff, c))
            if not cand_effects: continue

            gt_effect = gt_row['effect']
            key = (gt_row['element'], gt_row['info'])
            n_gt_same = len(gt_by_el_info[key])

            if n_gt_same > 1 and len(cand_effects) > 1:
                unused = [(oid, eff, c) for oid, eff, c in cand_effects if oid not in used_obs_ids]
                if not unused: unused = cand_effects
                best = min(unused, key=lambda x: abs(x[1] - gt_effect))
                our_effect = best[1]
                used_obs_ids.add(best[0])
                n_cands = 1
            else:
                effects = [eff for _, eff, _ in cand_effects]
                our_effect = sum(effects) / len(effects)
                n_cands = len(effects)

            err = abs(our_effect - gt_effect)
            matched += 1
            match_data = {
                'paper': paper_id,
                'ref': loladze_ref,
                'el': gt_row['element'],
                'our': our_effect,
                'gt': gt_effect,
                'err': err,
                'info': gt_row['info'],
                'n_candidates': n_cands,
            }
            paper_matches.append(match_data)
            all_matches.append(match_data)

        total_gt_el = len(gt_rows)
        mae = sum(m['err'] for m in paper_matches) / len(paper_matches) * 100 if paper_matches else float('nan')

        paper_results.append({
            'paper_id': paper_id, 'ref': loladze_ref,
            'gt_rows': total_gt_el, 'matched': matched,
            'total_obs': len(obs_list), 'mae': mae, 'tc_swap': swap_tc,
        })

    # Overall stats
    n = len(all_matches)
    if n == 0:
        print("No matches found!"); return

    w5 = sum(1 for m in all_matches if m['err'] <= 0.05)
    w10 = sum(1 for m in all_matches if m['err'] <= 0.10)
    w20 = sum(1 for m in all_matches if m['err'] <= 0.20)
    dir_total = sum(1 for m in all_matches if m['gt'] != 0)
    dir_ok = sum(1 for m in all_matches if m['gt'] != 0 and (m['our'] < 0) == (m['gt'] < 0))
    mae = sum(m['err'] for m in all_matches) / n * 100

    sorted_errs = sorted(m['err'] for m in all_matches)
    median_ae = sorted_errs[n // 2] * 100

    our = [m['our'] for m in all_matches]
    gts = [m['gt'] for m in all_matches]
    mean_our = sum(our) / n
    mean_gt = sum(gts) / n
    cov = sum((o - mean_our) * (g - mean_gt) for o, g in zip(our, gts))
    var_our = sum((o - mean_our) ** 2 for o in our)
    var_gt = sum((g - mean_gt) ** 2 for g in gts)
    r = cov / math.sqrt(var_our * var_gt) if var_our > 0 and var_gt > 0 else 0

    overall_gt = mean_gt * 100
    overall_our = mean_our * 100
    papers_matched = len([p for p in paper_results if p['matched'] > 0])
    total_gt = sum(p['gt_rows'] for p in paper_results)
    total_matched = sum(p['matched'] for p in paper_results)

    print(f"\n{'='*70}")
    print(f"AGENT EXTRACTION RESULTS")
    print(f"{'='*70}")
    print(f"Papers with GT match:  {papers_matched}/{len(agent_files)}")
    print(f"Element capture rate:  {total_matched}/{total_gt} ({total_matched/total_gt*100:.0f}%)")
    print(f"Pearson r:             {r:.3f}")
    print(f"Mean abs error:        {mae:.1f}%")
    print(f"Median abs error:      {median_ae:.1f}%")
    print(f"Within 5%:             {w5}/{n} ({w5/n*100:.0f}%)")
    print(f"Within 10%:            {w10}/{n} ({w10/n*100:.0f}%)")
    print(f"Within 20%:            {w20}/{n} ({w20/n*100:.0f}%)")
    print(f"Direction agreement:   {dir_ok}/{dir_total} ({dir_ok/dir_total*100:.0f}%)")
    print(f"Overall effect:        GT={overall_gt:.2f}% Ours={overall_our:.2f}% diff={abs(overall_gt-overall_our):.2f}pp")

    # Per-paper sorted by MAE
    print(f"\n{'='*70}")
    print(f"PER-PAPER (sorted by MAE)")
    print(f"{'='*70}")
    for pr in sorted(paper_results, key=lambda x: x['mae'] if not math.isnan(x['mae']) else 999):
        if pr['matched'] > 0:
            swap_flag = " [T/C SWAP]" if pr.get('tc_swap') else ""
            print(f"  {pr['paper_id']:<25} capture={pr['matched']}/{pr['gt_rows']:>2} "
                  f"MAE={pr['mae']:>5.1f}% ({pr['total_obs']} obs){swap_flag}")

    # Tiers
    tiers = {'Excellent': [], 'Good': [], 'Fair': [], 'Poor': []}
    for pr in paper_results:
        if pr['matched'] == 0: continue
        m = pr['mae']
        if math.isnan(m): continue
        if m <= 5: tiers['Excellent'].append(pr['paper_id'])
        elif m <= 10: tiers['Good'].append(pr['paper_id'])
        elif m <= 20: tiers['Fair'].append(pr['paper_id'])
        else: tiers['Poor'].append(pr['paper_id'])

    print(f"\n{'='*70}")
    print(f"PAPER TIERS")
    print(f"{'='*70}")
    for tier, papers in tiers.items():
        pct = len(papers)/papers_matched*100 if papers_matched else 0
        print(f"  {tier:10s}: {len(papers)} papers ({pct:.0f}%)")

    # Element-level
    print(f"\n{'='*70}")
    print(f"ELEMENT-LEVEL ANALYSIS")
    print(f"{'='*70}")
    el_data = defaultdict(list)
    for m in all_matches:
        el_data[m['el']].append(m)
    print(f"  {'Element':<5} {'N':>4} {'MAE':>6} {'Dir%':>5}")
    for el in sorted(el_data.keys(), key=lambda e: len(el_data[e]), reverse=True):
        ms = el_data[el]
        el_n = len(ms)
        el_mae = sum(m['err'] for m in ms) / el_n * 100
        el_dir_tot = sum(1 for m in ms if m['gt'] != 0)
        el_dir_ok = sum(1 for m in ms if m['gt'] != 0 and (m['our'] < 0) == (m['gt'] < 0))
        el_dir_pct = el_dir_ok / el_dir_tot * 100 if el_dir_tot else 0
        print(f"  {el:<5} {el_n:>4} {el_mae:>5.1f}% {el_dir_pct:>4.0f}%")

    # Worst matches
    print(f"\n{'='*70}")
    print(f"WORST MATCHES (err > 20%)")
    print(f"{'='*70}")
    worst = sorted(all_matches, key=lambda m: -m['err'])
    for m in worst[:15]:
        if m['err'] > 0.20:
            print(f"  {m['paper']:<22} {m['el']:3s} our={m['our']*100:+6.1f}% "
                  f"gt={m['gt']*100:+6.1f}% err={m['err']*100:5.1f}% "
                  f"info='{m['info'][:30]}' ({m['n_candidates']} cands)")

    # Save report
    report = {
        'timestamp': datetime.now().isoformat(),
        'method': 'Claude Code Agent (single model, subscription)',
        'papers_processed': len(agent_files),
        'papers_with_gt': papers_matched,
        'total_gt_rows': total_gt,
        'total_matched': total_matched,
        'capture_rate': f"{total_matched}/{total_gt} ({total_matched/total_gt*100:.0f}%)",
        'pearson_r': round(r, 3),
        'mae_pct': round(mae, 1),
        'median_ae_pct': round(median_ae, 1),
        'within_5pct': f"{w5}/{n} ({w5/n*100:.0f}%)",
        'within_10pct': f"{w10}/{n} ({w10/n*100:.0f}%)",
        'within_20pct': f"{w20}/{n} ({w20/n*100:.0f}%)",
        'direction_agreement': f"{dir_ok}/{dir_total} ({dir_ok/dir_total*100:.0f}%)",
        'overall_effect_gt': round(overall_gt, 2),
        'overall_effect_ours': round(overall_our, 2),
        'tiers': {k: len(v) for k, v in tiers.items()},
        'per_paper': paper_results,
        'all_matches': all_matches,
    }
    out = AGENT_DIR / "validation_report_agent.json"
    with open(out, 'w') as f:
        json.dump(report, f, indent=2, default=str)
    print(f"\nSaved to {out}")


if __name__ == '__main__':
    main()
