"""
Kimi-Primary + Fallback Simulation

Simulates a "Kimi-primary with Claude/Gemini fallback" architecture
using existing single-model baseline data, and compares it to the
current symmetric consensus pipeline.

Motivation:
  Ablation analysis shows Kimi solo (MAE=4.10%, r=0.903) outperforms
  consensus (MAE=4.54%, r=0.886) on a fixed scope of 322 observations.
  But consensus gives 560 total matched observations vs Kimi's ~486.
  A reviewer asks: why not just use Kimi as primary, and only invoke
  Claude/Gemini when Kimi fails?

Architecture simulated:
  For each paper:
    1. Use Kimi's extraction if it has >= 1 observation
    2. If Kimi has 0 observations, fall back to Claude
    3. If Claude also has 0, fall back to Gemini
    4. If all three have 0, the paper is skipped

Usage:
    .\\venv\\Scripts\\python.exe kimi_primary_simulation.py
"""
import sys
import os
import json
import math
import re
from pathlib import Path
from collections import defaultdict
from datetime import datetime

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv()

import openpyxl

# ==================================================================
# PATHS
# ==================================================================
BASE_DIR = Path(r"C:\Users\moshe\Dropbox\Testing metaanalyis program\meta_analysis_extractor")
GT_PATH = Path(r"C:\Users\moshe\Dropbox\Testing metaanalyis program\Loladze\CO2+Dataset.xlsx")

CONSENSUS_DIR = BASE_DIR / "output" / "loladze_full_46_v2"
BASELINE_DIRS = {
    "Kimi":   BASE_DIR / "output" / "baseline_kimi_loladze",
    "Claude": BASE_DIR / "output" / "baseline_claude_loladze",
    "Gemini": BASE_DIR / "output" / "baseline_gemini_loladze",
}

OUTPUT_DIR = BASE_DIR / "output"

# ==================================================================
# PAPER MAPPING (from validate_full_46.py)
# ==================================================================
PAPER_TO_LOLADZE_REF = {
    "001_Ma_2007": "Fernando et al 2012a",
    "002_Ziska_1997": "Ziska et al 1997",
    "003_Baslam_2012": "Baslam et al 2012",
    "004_Finzi_2001": "Finzi et al 2001",
    "005_Niinemets_1999": "Niinemets et al 1999",
    "006_Azam_2013": "Azam et al 2013",
    "007_Woodin_1992": "Woodin et al 1992",
    "008_Campbell_2002": "Cambell & Sage 2002",
    "009_Barnes_1992": "Barnes & Pffirrman 1992",
    "010_Li_2010": "Hogy et al 2009",
    "011_Huluka_1994": "Huluka et al 1994",
    "012_Wu_2004": "Wu et al 2004",
    "013_Keutgen_2001": "Keutgen & Chen 2001",
    "014_Lieffering_2004": "Lieffering et al 2004",
    "015_Pleijel_2009": "Pleijel & Danielsson 2009",
    "016_Fernando_2012a": "Fernando et al 2012b",
    "017_Fangmeier_2002": "Fangmeier et al  2002",
    "018_Al-Rawahy_2013": "Al-Rawahy et al 2013",
    "019_Baxter_1994": "Baxter et al 1994",
    "020_Overdieck_1993": "Overdieck 1993",
    "021_Wilsey_1994": "Wilsey et al 1994",
    "022_Blank_2011": "Blank et al 2011",
    "025_Guo_2011": "Guo et al 2013",
    "026_Seneweera_1997": "Seneweera & Conroy 1997",
    "027_Peet_1986": "Peet et al 1986",
    "028_Mishra_2011": "Mishra et al 2011",
    "031_Pal_2003": "Pal et al 2004",
    "032_Kanowski_2001": "Kanowski 2001",
    "034_Johnson_1997": "Johnson et al 2003",
    "035_Oksanen_2005": "Oksanen et al 2005",
    "036_Schenk_1997": "Schenk et al 1997",
    "037_de_2000": "Haase et al 2008",
    "038_Newbery_1995": "Newbery 1995",
    "039_Heagle_1993": "Heagle et al 1993",
    "040_Pfirrmann_1996": "Pfirrmann et al 1996",
    "041_Mjwara_1996": "Mjwara et al 1996",
    "042_Luomala_2005": "Luomala et al 2005",
    "043_Natali_2009": "Natali et al 2009",
    "044_Housman_2012": "Housman et al 2012",
    "046_Porter_1984": "Porter & Grodzinski 1984",
    "047_Rodenkirchen_2009": "Rodenkirchen et al 2009",
    "048_Khan_2013": "Khan et al 2012",
    "049_Singh_2013": "Singh et al 2013",
    "050_Polley_2011": "Polley et al 2011",
    "051_Niu_2013": "Niu et al 2013",
    "058_ONeill_1987": "O'Neill et al 1987",
    # Scanned papers added 2026-02-17
    "024_Nowak_2002": "Nowak et al 2002",
    "029_Kuehny_1991": "Kuehny et al 1991",
    "030_Wroblewitz_2013": "Wroblewitz et al 2013",
    "033_Johnson_2003": "Johnson et al 2003",
}

MISLABELED_PDFS = {
    "001_Ma_2007": "Fernando et al 2012a",
    "010_Li_2010": "Hogy et al 2009",
    "031_Pal_2003": "Pal et al 2004",
    "034_Johnson_1997": "Johnson et al 2003",
    "037_de_2000": "Haase et al 2008",
}


# ==================================================================
# HELPER FUNCTIONS (shared validation logic from validate_full_46.py)
# ==================================================================

def normalize_element(el):
    if not el:
        return ""
    el = el.strip()
    if '(' in el:
        el = el[:el.index('(')].strip()
    el = el.upper()
    MAP = {"NITROGEN": "N", "PHOSPHORUS": "P", "POTASSIUM": "K",
           "CALCIUM": "CA", "MAGNESIUM": "MG", "IRON": "FE",
           "ZINC": "ZN", "MANGANESE": "MN", "COPPER": "CU",
           "SULFUR": "S", "SULPHUR": "S", "BORON": "B", "CARBON": "C",
           "SODIUM": "NA", "MOLYBDENUM": "MO", "SILICON": "SI",
           "SELENIUM": "SE", "COBALT": "CO", "CHROMIUM": "CR",
           "CADMIUM": "CD", "NICKEL": "NI", "LEAD": "PB",
           "ALUMINIUM": "AL", "ALUMINUM": "AL", "BARIUM": "BA",
           "STRONTIUM": "SR", "BROMINE": "BR", "VANADIUM": "V",
           "CHLORINE": "CL"}
    return MAP.get(el, el)


def load_gt():
    """Load Loladze GT with Additional Info for each row."""
    wb = openpyxl.load_workbook(GT_PATH, data_only=True)
    ws = wb["CO2 Dataset"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[2]]

    col = {}
    for i, h in enumerate(headers):
        if h == 'Reference': col['ref'] = i
        elif h == '(E-A)/A': col['effect'] = i
        elif h == 'Element': col['element'] = i
        elif h == 'Additional Info': col['info'] = i
        elif h == 'eCO2': col['eco2'] = i
        elif h == 'aCO2': col['aco2'] = i
        elif h == 'Tissue': col['tissue'] = i
        elif h == 'Species': col['species'] = i
        elif h == 'n': col['n'] = i
        elif h == 'Foliar Edible': col['foliar_edible'] = i

    gt = defaultdict(list)
    for row in ws.iter_rows(min_row=3, values_only=True):
        ref = str(row[col['ref']]).strip() if row[col['ref']] else ""
        el = normalize_element(str(row[col['element']]).strip() if row[col['element']] else "")
        eff = row[col['effect']]
        info = str(row[col.get('info', 11)]).strip() if row[col.get('info', 11)] else ""
        if info == "None":
            info = ""
        eco2 = row[col.get('eco2', 8)]
        tissue = str(row[col.get('tissue', 4)]).strip() if row[col.get('tissue', 4)] else ""
        species = str(row[col.get('species', 0)]).strip() if col.get('species') is not None and row[col['species']] else ""
        n_val = row[col.get('n', 17)]

        if ref and el and eff is not None:
            try:
                gt[ref].append({
                    'element': el,
                    'effect': float(eff),
                    'info': info,
                    'eco2': float(eco2) if eco2 else None,
                    'tissue': tissue,
                    'species': species,
                    'n': int(n_val) if n_val else None,
                })
            except (ValueError, TypeError):
                pass

    wb.close()
    return dict(gt)


def get_mods(obs):
    mods = obs.get('moderators', {})
    if isinstance(mods, str):
        try:
            mods = json.loads(mods)
        except Exception:
            mods = {}
    if not isinstance(mods, dict):
        mods = {}
    return mods


def is_concentration_unit(unit_str):
    u = str(unit_str).lower()
    total_keywords = ['plant', 'pot', 'shoot', 'total', 'uptake', 'content']
    if any(kw in u for kw in total_keywords):
        return False
    return True


def filter_obs_for_gt_row(obs_list, gt_row, paper_id):
    """
    Filter observations to match a specific GT row's conditions.
    Same logic as ablation_analysis.py / validate_full_46.py.
    """
    gt_el = gt_row['element']
    gt_info = gt_row['info'].lower()

    el_matches = [o for o in obs_list if normalize_element(o.get('element', '')) == gt_el]
    if not el_matches:
        return []

    conc_matches = [o for o in el_matches if is_concentration_unit(o.get('unit', ''))]
    if conc_matches:
        el_matches = conc_matches

    if not gt_info:
        return el_matches

    filtered = el_matches

    # --- Leaf position filters (Baslam) ---
    if 'inner' in gt_info and 'nm' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            leaf_pos = str(mods.get('leaf_position', '')).lower()
            amf = str(mods.get('AMF_status', mods.get('mycorrhizal_status', ''))).lower()
            if 'inner' in leaf_pos and ('non' in amf or 'nm' in amf):
                new.append(o)
        if new:
            filtered = new
    elif 'nm' in gt_info or 'non-mycorrhizal' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            amf = str(mods.get('AMF_status', mods.get('mycorrhizal_status', ''))).lower()
            if 'non' in amf or 'nm' in amf:
                new.append(o)
        if new:
            filtered = new

    # --- Leaf age ---
    if 'old leaves' in gt_info:
        new = [o for o in filtered if 'old' in str(get_mods(o).get('leaf_age', '')).lower()]
        if new: filtered = new
    if 'young leaves' in gt_info:
        new = [o for o in filtered if 'young' in str(get_mods(o).get('leaf_age', '')).lower()]
        if new: filtered = new

    # --- Year filter ---
    year_match = re.match(r'^(\d{4})$', gt_row['info'].strip())
    if year_match:
        target_year = year_match.group(1)
        new = []
        for o in filtered:
            mods = get_mods(o)
            if target_year in str(mods.get('year', '')):
                new.append(o)
        if new:
            filtered = new

    # --- High/Low N ---
    if 'high n' in gt_info and 'low' not in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            n_level = str(mods.get('nitrogen_level', mods.get('N_level', mods.get('n_treatment', '')))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if 'high' in n_level or 'high n' in desc:
                new.append(o)
        if new: filtered = new
    elif 'low n' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            n_level = str(mods.get('nitrogen_level', mods.get('N_level', mods.get('n_treatment', '')))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if 'low' in n_level or 'low n' in desc:
                new.append(o)
        if new: filtered = new

    # --- Soil type ---
    if 'basalt' in gt_info:
        new = [o for o in filtered
               if 'basalt' in str(get_mods(o).get('soil_type', '')).lower()
               or 'basalt' in str(o.get('treatment_description', '')).lower()]
        if new: filtered = new
    if 'rhyolite' in gt_info:
        new = [o for o in filtered
               if 'rhyolite' in str(get_mods(o).get('soil_type', '')).lower()
               or 'rhyolite' in str(o.get('treatment_description', '')).lower()]
        if new: filtered = new

    # --- Cultivar (Heagle) ---
    for cultivar_name in ['nc-r', 'nc-s']:
        if cultivar_name in gt_info:
            new = []
            for o in filtered:
                mods = get_mods(o)
                cult = str(mods.get('cultivar', mods.get('clone', ''))).lower()
                desc = str(o.get('treatment_description', '')).lower()
                if cultivar_name in cult or cultivar_name in desc:
                    new.append(o)
            if new: filtered = new
            break

    # --- Potassium treatment (Pfirrmann) ---
    if '+k' in gt_info and '-k' not in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            k_treat = str(mods.get('potassium', mods.get('K_treatment', ''))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if '+k' in k_treat or 'plus k' in k_treat or '+k' in desc or 'with k' in desc:
                new.append(o)
        if new: filtered = new
    elif '-k' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            k_treat = str(mods.get('potassium', mods.get('K_treatment', ''))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if '-k' in k_treat or 'minus k' in k_treat or '-k' in desc or 'without k' in desc:
                new.append(o)
        if new: filtered = new

    # --- P level (Seneweera) ---
    p_level_match = re.search(r'p\s+(\d+)\s*mg/kg', gt_info)
    if p_level_match:
        target_p = p_level_match.group(1)
        target_p_int = int(target_p)
        new = []
        for o in filtered:
            mods = get_mods(o)
            all_mods_str = str(mods).lower()
            desc = str(o.get('treatment_description', '')).lower()
            for mk, mv in mods.items():
                mk_lower = str(mk).lower()
                if ('p_level' in mk_lower or 'phosph' in mk_lower) and mv is not None:
                    try:
                        if abs(float(mv) - target_p_int) < 1:
                            new.append(o)
                            break
                    except (ValueError, TypeError):
                        if target_p in str(mv):
                            new.append(o)
                            break
            else:
                if f'p {target_p}' in desc or f'{target_p} mg' in all_mods_str:
                    new.append(o)
        if new: filtered = new

    # --- Harvest year ---
    harvest_year_match = re.search(r'(\d{4})\s*harvest', gt_info)
    if harvest_year_match:
        target_year = harvest_year_match.group(1)
        new = []
        for o in filtered:
            mods = get_mods(o)
            all_text = str(mods) + ' ' + str(o.get('treatment_description', ''))
            if target_year in all_text:
                new.append(o)
        if new: filtered = new

    # --- Natali site filters ---
    if paper_id == '043_Natali_2009':
        if 'duke' in gt_info and 'serc' not in gt_info:
            new = [o for o in filtered
                   if 'duke' in str(get_mods(o).get('site', '')).lower()
                   or 'duke' in str(o.get('treatment_description', '')).lower()]
            if new: filtered = new
            if '1 yr old' in gt_info:
                new = [o for o in filtered
                       if 'pinus' in str(get_mods(o).get('species', '')).lower()
                       or 'taeda' in str(get_mods(o).get('species', '')).lower()]
                if new: filtered = new
            else:
                new = [o for o in filtered
                       if 'liquidambar' in str(get_mods(o).get('species', '')).lower()
                       or 'sweetgum' in str(o.get('treatment_description', '')).lower()]
                if new: filtered = new
        elif 'ornl' in gt_info:
            new = [o for o in filtered
                   if 'ornl' in str(get_mods(o).get('site', '')).lower()]
            if new: filtered = new
        elif 'serc' in gt_info:
            new = [o for o in filtered
                   if 'serc' in str(get_mods(o).get('site', '')).lower()]
            if new: filtered = new
            gt_species = gt_row.get('species', '').lower()
            if gt_species:
                gt_parts = [p for p in gt_species.split() if len(p) > 3]
                species_new = []
                for o in filtered:
                    obs_sp = str(get_mods(o).get('species', '')).lower()
                    desc = str(o.get('treatment_description', '')).lower()
                    combined = obs_sp + ' ' + desc
                    if all(part in combined for part in gt_parts):
                        species_new.append(o)
                if species_new: filtered = species_new

    # --- Ozone filter ---
    if 'o3' not in gt_info and 'ozone' not in gt_info:
        has_o3_mods = any(
            any(k.lower() in ('o3_level', 'ozone', 'o3', 'ozone_level')
                for k in get_mods(o).keys())
            for o in filtered
        )
        if has_o3_mods:
            ambient_o3 = []
            for o in filtered:
                mods = get_mods(o)
                o3_val = ''
                for k, v in mods.items():
                    if k.lower() in ('o3_level', 'ozone', 'o3', 'ozone_level'):
                        o3_val = str(v).lower()
                        break
                if any(term in o3_val for term in ['ambient', 'cf', 'charcoal',
                       'control', 'low', 'clean', 'filtered', '5 nl',
                       '20 n', '20n', 'background']):
                    ambient_o3.append(o)
                elif not o3_val:
                    ambient_o3.append(o)
            if ambient_o3:
                filtered = ambient_o3

    # --- CO2 level filter ---
    if gt_row.get('eco2'):
        gt_co2 = gt_row['eco2']
        new = []
        for o in filtered:
            desc = str(o.get('treatment_description', '')).lower()
            mods = get_mods(o)
            co2_match_re = re.search(r'(\d{3,4})\s*(?:ppm|umol|umol|mumol)', desc)
            our_co2 = None
            if co2_match_re:
                our_co2 = float(co2_match_re.group(1))
            else:
                for k, v in mods.items():
                    k_l = k.lower()
                    if 'co2' in k_l and v is not None:
                        try:
                            our_co2 = float(v)
                            break
                        except (ValueError, TypeError):
                            co2_in_mod = re.search(r'(\d{3,4})', str(v))
                            if co2_in_mod:
                                our_co2 = float(co2_in_mod.group(1))
                                break
            if our_co2 is not None:
                if abs(our_co2 - gt_co2) < 100:
                    new.append(o)
            else:
                new.append(o)
        if new:
            filtered = new

    return filtered


def deduplicate_vision_text(obs_list):
    """Prefer text over vision when both exist."""
    text_obs = []
    vision_obs = []
    for o in obs_list:
        notes = str(o.get('notes', '')).lower()
        if '[from vision]' in notes or 'vision' in notes:
            vision_obs.append(o)
        else:
            text_obs.append(o)
    if not vision_obs:
        return obs_list
    text_keys = set()
    for o in text_obs:
        mods = get_mods(o)
        el = normalize_element(o.get('element', ''))
        sp = str(mods.get('species', '')).lower().strip()
        site = str(mods.get('site', '')).lower().strip()
        text_keys.add((el, sp, site))
    deduped = list(text_obs)
    for o in vision_obs:
        mods = get_mods(o)
        el = normalize_element(o.get('element', ''))
        sp = str(mods.get('species', '')).lower().strip()
        site = str(mods.get('site', '')).lower().strip()
        if (el, sp, site) not in text_keys:
            deduped.append(o)
    return deduped


def detect_tc_swap(obs_list, gt_rows):
    """Detect if extraction has treatment/control swapped."""
    our_effects = []
    gt_effects = []
    for gt_row in gt_rows:
        gt_el = gt_row['element']
        matching = [o for o in obs_list
                    if normalize_element(o.get('element', '')) == gt_el
                    and is_concentration_unit(o.get('unit', ''))]
        if not matching:
            continue
        for o in matching:
            ctrl = o.get('control_mean')
            treat = o.get('treatment_mean')
            if ctrl and ctrl != 0 and treat is not None:
                our_eff = (treat - ctrl) / ctrl
                if abs(our_eff) <= 5.0:
                    our_effects.append(our_eff)
                    gt_effects.append(gt_row['effect'])

    if len(our_effects) < 3:
        return False

    disagree = sum(1 for o, g in zip(our_effects, gt_effects)
                   if g != 0 and (o < 0) != (g < 0))
    return disagree / len(our_effects) > 0.7


def compute_effect(obs, swap_tc=False):
    ctrl = obs.get('control_mean')
    treat = obs.get('treatment_mean')
    if swap_tc:
        ctrl, treat = treat, ctrl
    if ctrl and ctrl != 0 and treat is not None:
        return (treat - ctrl) / ctrl
    return None


# ==================================================================
# CORE VALIDATION
# ==================================================================

def validate_obs_list(obs_list, gt_rows, paper_id):
    """
    Validate a list of observations for one paper against GT rows.
    Returns list of match dicts.
    """
    if not obs_list:
        return []

    # Deduplicate vision/text
    obs_list = deduplicate_vision_text(obs_list)

    # Filter to concentration units
    has_conc = any(is_concentration_unit(o.get('unit', '')) for o in obs_list)
    has_total = any(not is_concentration_unit(o.get('unit', ''))
                    for o in obs_list if o.get('unit', ''))
    if not has_conc and has_total:
        return []

    # Filter sub-ambient CO2
    elevated_obs = []
    for o in obs_list:
        desc = str(o.get('treatment_description', '')).lower()
        co2_match = re.search(r'(\d{2,4})\s*(?:ppm|umol|umol|mumol)', desc)
        if co2_match:
            co2_val = float(co2_match.group(1))
            if co2_val < 300:
                continue
        if 'low co2' in desc or 'sub-ambient' in desc:
            continue
        elevated_obs.append(o)
    if elevated_obs:
        obs_list = elevated_obs

    # Detect T/C swap
    swap_tc = detect_tc_swap(obs_list, gt_rows)

    # Match GT rows
    gt_by_el_info = defaultdict(list)
    for gt_row in gt_rows:
        key = (gt_row['element'], gt_row['info'])
        gt_by_el_info[key].append(gt_row)

    used_obs_ids = set()
    matches = []

    for gt_row in gt_rows:
        candidates = filter_obs_for_gt_row(obs_list, gt_row, paper_id)
        if not candidates:
            continue

        cand_effects = []
        for c in candidates:
            obs_id = id(c)
            eff = compute_effect(c, swap_tc=swap_tc)
            if eff is not None and abs(eff) <= 5.0:
                cand_effects.append((obs_id, eff, c))

        if not cand_effects:
            continue

        gt_effect = gt_row['effect']
        key = (gt_row['element'], gt_row['info'])
        n_gt_same = len(gt_by_el_info[key])

        if n_gt_same > 1 and len(cand_effects) > 1:
            unused = [(oid, eff, c) for oid, eff, c in cand_effects
                      if oid not in used_obs_ids]
            if not unused:
                unused = cand_effects
            best = min(unused, key=lambda x: abs(x[1] - gt_effect))
            our_effect = best[1]
            used_obs_ids.add(best[0])
        else:
            effects = [eff for _, eff, _ in cand_effects]
            our_effect = sum(effects) / len(effects)

        err = abs(our_effect - gt_effect)
        matches.append({
            'paper': paper_id,
            'el': gt_row['element'],
            'our': our_effect,
            'gt': gt_effect,
            'err': err,
            'info': gt_row['info'],
        })

    return matches


def validate_source(source_dir, gt, source_type="consensus"):
    """Validate all papers in a source directory against GT."""
    if source_type == "consensus":
        pattern = "*_consensus.json"
        obs_key = "consensus_observations"
    else:
        pattern = "*_baseline.json"
        obs_key = "observations"

    results_files = sorted(source_dir.glob(pattern))
    all_matches = []
    paper_details = {}

    for rf in results_files:
        if rf.name.startswith('_'):
            continue

        paper_id = rf.stem.replace("_consensus", "").replace("_baseline", "")
        loladze_ref = PAPER_TO_LOLADZE_REF.get(paper_id)

        if not loladze_ref or loladze_ref not in gt:
            for ref in gt:
                surname = paper_id.split('_')[1] if '_' in paper_id else paper_id
                if surname.lower() in ref.lower():
                    loladze_ref = ref
                    break
            if not loladze_ref or loladze_ref not in gt:
                continue

        with open(rf, encoding='utf-8') as f:
            data = json.load(f)

        gt_rows = gt[loladze_ref]
        obs_list = data.get(obs_key, [])

        matches = validate_obs_list(obs_list, gt_rows, paper_id)
        all_matches.extend(matches)

        paper_details[paper_id] = {
            'gt_rows': len(gt_rows),
            'matched': len(matches),
            'total_obs': len(obs_list),
        }

    return all_matches, paper_details


def compute_stats(matches):
    """Compute aggregate statistics from a list of match dicts."""
    n = len(matches)
    if n == 0:
        return {
            'n': 0, 'mae': float('nan'), 'median_ae': float('nan'),
            'r': float('nan'), 'direction_pct': float('nan'),
            'w5': 0, 'w10': 0, 'w20': 0,
            'overall_our': float('nan'), 'overall_gt': float('nan'),
            'papers_matched': 0,
        }

    errs = [m['err'] for m in matches]
    mae = sum(errs) / n * 100
    sorted_errs = sorted(errs)
    median_ae = sorted_errs[n // 2] * 100

    w5 = sum(1 for e in errs if e <= 0.05)
    w10 = sum(1 for e in errs if e <= 0.10)
    w20 = sum(1 for e in errs if e <= 0.20)

    our = [m['our'] for m in matches]
    gts = [m['gt'] for m in matches]

    dir_total = sum(1 for m in matches if m['gt'] != 0)
    dir_ok = sum(1 for m in matches if m['gt'] != 0 and (m['our'] < 0) == (m['gt'] < 0))
    dir_pct = dir_ok / dir_total * 100 if dir_total > 0 else 0

    mean_our = sum(our) / n
    mean_gt = sum(gts) / n
    cov = sum((o - mean_our) * (g - mean_gt) for o, g in zip(our, gts))
    var_our = sum((o - mean_our) ** 2 for o in our)
    var_gt = sum((g - mean_gt) ** 2 for g in gts)
    r = cov / math.sqrt(var_our * var_gt) if var_our > 0 and var_gt > 0 else 0

    papers_matched = len(set(m['paper'] for m in matches))

    return {
        'n': n,
        'mae': round(mae, 2),
        'median_ae': round(median_ae, 2),
        'r': round(r, 3),
        'direction_pct': round(dir_pct, 1),
        'w5': w5,
        'w10': w10,
        'w20': w20,
        'w5_pct': round(w5 / n * 100, 1),
        'w10_pct': round(w10 / n * 100, 1),
        'w20_pct': round(w20 / n * 100, 1),
        'overall_our': round(mean_our * 100, 2),
        'overall_gt': round(mean_gt * 100, 2),
        'papers_matched': papers_matched,
    }


# ==================================================================
# KIMI-PRIMARY SIMULATION
# ==================================================================

def load_baseline_obs_by_paper(baseline_dir):
    """
    Load all per-paper observation lists from a baseline directory.
    Returns dict: paper_id -> list of observations.
    """
    obs_by_paper = {}
    for rf in sorted(baseline_dir.glob("*_baseline.json")):
        if rf.name.startswith('_'):
            continue
        paper_id = rf.stem.replace("_baseline", "")
        with open(rf, encoding='utf-8') as f:
            data = json.load(f)
        obs_by_paper[paper_id] = data.get('observations', [])
    return obs_by_paper


def simulate_kimi_primary(gt):
    """
    Simulate a Kimi-primary + fallback architecture:
      - For each paper: use Kimi's extraction if it has >= 1 observation
      - If Kimi has 0, fall back to Claude
      - If Claude also has 0, fall back to Gemini

    Returns (all_matches, paper_details, routing_log).
    """
    # Load all baselines
    kimi_obs = load_baseline_obs_by_paper(BASELINE_DIRS["Kimi"])
    claude_obs = load_baseline_obs_by_paper(BASELINE_DIRS["Claude"])
    gemini_obs = load_baseline_obs_by_paper(BASELINE_DIRS["Gemini"])

    # Collect all paper IDs across all sources
    all_paper_ids = sorted(set(kimi_obs.keys()) | set(claude_obs.keys()) | set(gemini_obs.keys()))

    all_matches = []
    paper_details = {}
    routing_log = []

    for paper_id in all_paper_ids:
        loladze_ref = PAPER_TO_LOLADZE_REF.get(paper_id)
        if not loladze_ref or loladze_ref not in gt:
            # Try fuzzy matching
            for ref in gt:
                surname = paper_id.split('_')[1] if '_' in paper_id else paper_id
                if surname.lower() in ref.lower():
                    loladze_ref = ref
                    break
            if not loladze_ref or loladze_ref not in gt:
                continue

        gt_rows = gt[loladze_ref]

        # Kimi-primary fallback logic
        k_obs = kimi_obs.get(paper_id, [])
        c_obs = claude_obs.get(paper_id, [])
        g_obs = gemini_obs.get(paper_id, [])

        if len(k_obs) > 0:
            selected_obs = k_obs
            source_used = "Kimi"
        elif len(c_obs) > 0:
            selected_obs = c_obs
            source_used = "Claude (fallback)"
        elif len(g_obs) > 0:
            selected_obs = g_obs
            source_used = "Gemini (fallback)"
        else:
            selected_obs = []
            source_used = "NONE"

        routing_log.append({
            'paper_id': paper_id,
            'kimi_obs': len(k_obs),
            'claude_obs': len(c_obs),
            'gemini_obs': len(g_obs),
            'source_used': source_used,
            'obs_used': len(selected_obs),
        })

        if not selected_obs:
            continue

        matches = validate_obs_list(selected_obs, gt_rows, paper_id)
        all_matches.extend(matches)

        paper_details[paper_id] = {
            'gt_rows': len(gt_rows),
            'matched': len(matches),
            'total_obs': len(selected_obs),
            'source_used': source_used,
        }

    return all_matches, paper_details, routing_log


def simulate_kimi_primary_obs_level(gt):
    """
    More granular Kimi-primary simulation:
    For each paper, use Kimi's extraction. But if Kimi's extraction
    yields 0 GT MATCHES (not just 0 obs), fall back to Claude, then Gemini.

    This handles the case where Kimi extracts data but it doesn't match
    any GT rows (e.g., wrong elements or wrong tissue type).
    """
    kimi_obs = load_baseline_obs_by_paper(BASELINE_DIRS["Kimi"])
    claude_obs = load_baseline_obs_by_paper(BASELINE_DIRS["Claude"])
    gemini_obs = load_baseline_obs_by_paper(BASELINE_DIRS["Gemini"])

    all_paper_ids = sorted(set(kimi_obs.keys()) | set(claude_obs.keys()) | set(gemini_obs.keys()))

    all_matches = []
    paper_details = {}
    routing_log = []

    for paper_id in all_paper_ids:
        loladze_ref = PAPER_TO_LOLADZE_REF.get(paper_id)
        if not loladze_ref or loladze_ref not in gt:
            for ref in gt:
                surname = paper_id.split('_')[1] if '_' in paper_id else paper_id
                if surname.lower() in ref.lower():
                    loladze_ref = ref
                    break
            if not loladze_ref or loladze_ref not in gt:
                continue

        gt_rows = gt[loladze_ref]

        # Try Kimi first, fall back on zero GT matches
        k_obs = kimi_obs.get(paper_id, [])
        c_obs = claude_obs.get(paper_id, [])
        g_obs = gemini_obs.get(paper_id, [])

        candidates = [
            ("Kimi", k_obs),
            ("Claude (fallback)", c_obs),
            ("Gemini (fallback)", g_obs),
        ]

        best_matches = []
        source_used = "NONE"

        for src_name, src_obs in candidates:
            if len(src_obs) == 0:
                continue
            matches = validate_obs_list(src_obs, gt_rows, paper_id)
            if len(matches) > 0:
                best_matches = matches
                source_used = src_name
                break

        routing_log.append({
            'paper_id': paper_id,
            'kimi_obs': len(k_obs),
            'claude_obs': len(c_obs),
            'gemini_obs': len(g_obs),
            'source_used': source_used,
            'gt_matches': len(best_matches),
        })

        if best_matches:
            all_matches.extend(best_matches)
            paper_details[paper_id] = {
                'gt_rows': len(gt_rows),
                'matched': len(best_matches),
                'source_used': source_used,
            }

    return all_matches, paper_details, routing_log


# ==================================================================
# MAIN
# ==================================================================

def main():
    print("=" * 78)
    print("KIMI-PRIMARY + FALLBACK SIMULATION")
    print("Simulating: Use Kimi as primary, Claude/Gemini as fallback")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 78)

    # Load ground truth
    gt = load_gt()
    total_gt_rows = sum(len(v) for v in gt.values())
    print(f"\nGround truth: {total_gt_rows} observations across {len(gt)} references\n")

    # ------------------------------------------------------------------
    # 1. Validate existing sources (baselines)
    # ------------------------------------------------------------------
    print("Validating existing sources...")

    print("  [1/4] Consensus pipeline...")
    consensus_matches, consensus_papers = validate_source(CONSENSUS_DIR, gt, "consensus")
    consensus_stats = compute_stats(consensus_matches)

    print("  [2/4] Kimi solo...")
    kimi_matches, kimi_papers = validate_source(BASELINE_DIRS["Kimi"], gt, "baseline")
    kimi_stats = compute_stats(kimi_matches)

    print("  [3/4] Claude solo...")
    claude_matches, claude_papers = validate_source(BASELINE_DIRS["Claude"], gt, "baseline")
    claude_stats = compute_stats(claude_matches)

    print("  [4/4] Gemini solo...")
    gemini_matches, gemini_papers = validate_source(BASELINE_DIRS["Gemini"], gt, "baseline")
    gemini_stats = compute_stats(gemini_matches)

    # ------------------------------------------------------------------
    # 2. Simulate Kimi-primary + fallback (obs-level)
    # ------------------------------------------------------------------
    print("\n  Simulating Kimi-primary + obs-level fallback...")
    kp_obs_matches, kp_obs_papers, kp_obs_routing = simulate_kimi_primary_obs_level(gt)
    kp_obs_stats = compute_stats(kp_obs_matches)

    # ------------------------------------------------------------------
    # 3. Simulate Kimi-primary + fallback (paper-level)
    # ------------------------------------------------------------------
    print("  Simulating Kimi-primary + paper-level fallback...")
    kp_matches, kp_papers, kp_routing = simulate_kimi_primary(gt)
    kp_stats = compute_stats(kp_matches)

    # ------------------------------------------------------------------
    # RESULTS
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("TABLE 1: FULL-SCOPE COMPARISON (all matched observations)")
    print(f"{'='*78}")
    print(f"{'Method':<32} {'N obs':>6} {'Papers':>7} {'MAE%':>7} {'r':>7} {'Dir%':>6} {'OverallEff':>11}")
    print("-" * 78)

    rows = [
        ("Consensus (current)", consensus_stats),
        ("Kimi-primary + fallback (paper)", kp_stats),
        ("Kimi-primary + fallback (obs)", kp_obs_stats),
        ("Kimi solo (no fallback)", kimi_stats),
        ("Claude solo", claude_stats),
        ("Gemini solo", gemini_stats),
    ]

    for label, s in rows:
        if s['n'] > 0:
            overall_diff = abs(s['overall_our'] - s['overall_gt'])
            print(f"{label:<32} {s['n']:>6} {s['papers_matched']:>7} "
                  f"{s['mae']:>6.2f}% {s['r']:>6.3f} {s['direction_pct']:>5.1f}% "
                  f"{s['overall_our']:>+6.2f}%")
        else:
            print(f"{label:<32} {'N/A':>6}")

    # GT overall effect for reference
    all_gt_effects = []
    for ref_rows in gt.values():
        for row in ref_rows:
            all_gt_effects.append(row['effect'])
    gt_overall = sum(all_gt_effects) / len(all_gt_effects) * 100 if all_gt_effects else 0
    print(f"\n  GT overall effect (full dataset): {gt_overall:+.2f}%")
    print(f"  GT overall effect (consensus scope): {consensus_stats['overall_gt']:+.2f}%")

    # ------------------------------------------------------------------
    # ROUTING ANALYSIS
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("TABLE 2: ROUTING ANALYSIS (paper-level fallback)")
    print(f"{'='*78}")

    kimi_count = sum(1 for r in kp_routing if r['source_used'] == 'Kimi')
    claude_fb = sum(1 for r in kp_routing if r['source_used'] == 'Claude (fallback)')
    gemini_fb = sum(1 for r in kp_routing if r['source_used'] == 'Gemini (fallback)')
    none_count = sum(1 for r in kp_routing if r['source_used'] == 'NONE')

    print(f"  Kimi primary:      {kimi_count} papers")
    print(f"  Claude fallback:   {claude_fb} papers")
    print(f"  Gemini fallback:   {gemini_fb} papers")
    print(f"  No data (all 0):   {none_count} papers")
    print(f"  Total:             {len(kp_routing)} papers")

    # Show which papers fell back
    print(f"\n  Papers where fallback was needed:")
    for r in kp_routing:
        if 'fallback' in r['source_used']:
            print(f"    {r['paper_id']:<25} -> {r['source_used']:<20} "
                  f"(Kimi: {r['kimi_obs']} obs, Claude: {r['claude_obs']} obs, "
                  f"Gemini: {r['gemini_obs']} obs)")

    print(f"\n  Papers with no data from any model:")
    for r in kp_routing:
        if r['source_used'] == 'NONE':
            print(f"    {r['paper_id']}")

    # ------------------------------------------------------------------
    # OBS-LEVEL ROUTING ANALYSIS
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("TABLE 3: ROUTING ANALYSIS (obs-level fallback)")
    print(f"{'='*78}")

    kimi_count_obs = sum(1 for r in kp_obs_routing if r['source_used'] == 'Kimi')
    claude_fb_obs = sum(1 for r in kp_obs_routing if r['source_used'] == 'Claude (fallback)')
    gemini_fb_obs = sum(1 for r in kp_obs_routing if r['source_used'] == 'Gemini (fallback)')
    none_count_obs = sum(1 for r in kp_obs_routing if r['source_used'] == 'NONE')

    print(f"  Kimi primary:      {kimi_count_obs} papers")
    print(f"  Claude fallback:   {claude_fb_obs} papers")
    print(f"  Gemini fallback:   {gemini_fb_obs} papers")
    print(f"  No GT match:       {none_count_obs} papers")

    print(f"\n  Papers where obs-level fallback kicked in:")
    for r in kp_obs_routing:
        if 'fallback' in r['source_used']:
            print(f"    {r['paper_id']:<25} -> {r['source_used']:<20} "
                  f"(Kimi: {r['kimi_obs']} obs, GT matches: {r['gt_matches']})")

    # ------------------------------------------------------------------
    # FIXED-SCOPE COMPARISON
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("TABLE 4: FIXED-SCOPE COMPARISON (same GT rows for all methods)")
    print(f"{'='*78}")

    def match_key(m):
        return (m['paper'], m['el'], m['info'])

    consensus_keys = set(match_key(m) for m in consensus_matches)
    kimi_keys = set(match_key(m) for m in kimi_matches)
    kp_keys = set(match_key(m) for m in kp_matches)
    kp_obs_keys = set(match_key(m) for m in kp_obs_matches)

    # Common to all four
    common_keys = consensus_keys & kimi_keys & kp_keys & kp_obs_keys

    print(f"\n  GT rows matched by all 4 methods: {len(common_keys)}")

    if common_keys:
        fixed_consensus = [m for m in consensus_matches if match_key(m) in common_keys]
        fixed_kimi = [m for m in kimi_matches if match_key(m) in common_keys]
        fixed_kp = [m for m in kp_matches if match_key(m) in common_keys]
        fixed_kp_obs = [m for m in kp_obs_matches if match_key(m) in common_keys]

        f_consensus = compute_stats(fixed_consensus)
        f_kimi = compute_stats(fixed_kimi)
        f_kp = compute_stats(fixed_kp)
        f_kp_obs = compute_stats(fixed_kp_obs)

        print(f"\n  {'Method':<32} {'N obs':>6} {'MAE%':>7} {'r':>7} {'Dir%':>6}")
        print("  " + "-" * 60)

        for label, s in [
            ("Consensus (current)", f_consensus),
            ("Kimi-primary + fallback (paper)", f_kp),
            ("Kimi-primary + fallback (obs)", f_kp_obs),
            ("Kimi solo (no fallback)", f_kimi),
        ]:
            print(f"  {label:<32} {s['n']:>6} {s['mae']:>6.2f}% {s['r']:>6.3f} {s['direction_pct']:>5.1f}%")

    # ------------------------------------------------------------------
    # COVERAGE vs ACCURACY TRADEOFF
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("TABLE 5: COVERAGE vs ACCURACY TRADEOFF")
    print(f"{'='*78}")

    # Total possible GT rows for our papers
    all_paper_gt_rows = set()
    for paper_id in PAPER_TO_LOLADZE_REF:
        ref = PAPER_TO_LOLADZE_REF[paper_id]
        if ref in gt:
            for i, row in enumerate(gt[ref]):
                all_paper_gt_rows.add((paper_id, row['element'], row['info']))
    total_possible = len(all_paper_gt_rows)

    print(f"\n  Total GT observations for our papers: {total_possible}")
    print()
    print(f"  {'Method':<32} {'Matched':>8} {'Cover%':>8} {'MAE%':>7} {'r':>7} {'Dir%':>6}")
    print("  " + "-" * 70)

    for label, s in rows:
        if s['n'] > 0:
            cov = s['n'] / total_possible * 100 if total_possible > 0 else 0
            print(f"  {label:<32} {s['n']:>8} {cov:>7.1f}% {s['mae']:>6.2f}% {s['r']:>6.3f} {s['direction_pct']:>5.1f}%")

    # ------------------------------------------------------------------
    # KEY FINDING: WHERE DOES KIMI-PRIMARY DIFFER FROM CONSENSUS?
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("TABLE 6: PER-PAPER COMPARISON (Kimi-primary vs Consensus)")
    print(f"{'='*78}")

    # For each paper, compare the match count
    cons_paper_matches = defaultdict(list)
    for m in consensus_matches:
        cons_paper_matches[m['paper']].append(m)

    kp_paper_matches = defaultdict(list)
    for m in kp_obs_matches:
        kp_paper_matches[m['paper']].append(m)

    all_papers_seen = sorted(set(list(cons_paper_matches.keys()) + list(kp_paper_matches.keys())))

    better_consensus = []
    better_kp = []
    tied = []

    print(f"\n  {'Paper':<25} {'Cons N':>7} {'ConMAE':>7} {'KP N':>6} {'KP MAE':>7} {'KP src':>20} {'Winner':>10}")
    print("  " + "-" * 85)

    for paper_id in all_papers_seen:
        c_matches = cons_paper_matches.get(paper_id, [])
        k_matches = kp_paper_matches.get(paper_id, [])

        c_n = len(c_matches)
        k_n = len(k_matches)
        c_mae = sum(m['err'] for m in c_matches) / c_n * 100 if c_n > 0 else float('nan')
        k_mae = sum(m['err'] for m in k_matches) / k_n * 100 if k_n > 0 else float('nan')

        # What source did Kimi-primary use?
        kp_src = ""
        for r in kp_obs_routing:
            if r['paper_id'] == paper_id:
                kp_src = r['source_used']
                break

        # Who wins on (coverage, then MAE)?
        if c_n > k_n:
            winner = "Consensus"
            better_consensus.append(paper_id)
        elif k_n > c_n:
            winner = "Kimi-primary"
            better_kp.append(paper_id)
        else:
            if not math.isnan(c_mae) and not math.isnan(k_mae):
                if c_mae < k_mae - 0.5:
                    winner = "Consensus"
                    better_consensus.append(paper_id)
                elif k_mae < c_mae - 0.5:
                    winner = "Kimi-primary"
                    better_kp.append(paper_id)
                else:
                    winner = "Tied"
                    tied.append(paper_id)
            else:
                winner = "Tied"
                tied.append(paper_id)

        # Only show papers where they differ
        if c_n != k_n or (not math.isnan(c_mae) and not math.isnan(k_mae) and abs(c_mae - k_mae) > 0.5):
            print(f"  {paper_id:<25} {c_n:>7} {c_mae:>6.1f}% {k_n:>6} {k_mae:>6.1f}% {kp_src:>20} {winner:>10}")

    print(f"\n  Summary: Consensus better in {len(better_consensus)} papers, "
          f"Kimi-primary better in {len(better_kp)}, tied in {len(tied)}")

    # ------------------------------------------------------------------
    # COST ANALYSIS
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("TABLE 7: COST ANALYSIS (estimated)")
    print(f"{'='*78}")

    n_papers = len(kp_routing)
    n_kimi_calls = kimi_count  # from paper-level routing
    n_fallback_calls = claude_fb + gemini_fb

    print(f"\n  Consensus (current architecture):")
    print(f"    Kimi calls:    {n_papers} papers x 1 = {n_papers}")
    print(f"    Claude calls:  {n_papers} papers x 1 = {n_papers}")
    print(f"    Gemini calls:  ~{n_papers // 4} papers (tiebreaker, ~25%)")
    print(f"    Total API calls: ~{n_papers * 2 + n_papers // 4}")
    print(f"    Estimated cost: ~$17 (as reported)")

    print(f"\n  Kimi-primary + fallback:")
    print(f"    Kimi calls:    {n_papers} papers (always called first)")
    print(f"    Fallback calls: {n_fallback_calls} papers (Claude or Gemini)")
    print(f"    Total API calls: {n_papers + n_fallback_calls}")

    # Estimated cost per model per paper
    kimi_cost_per = 0.03
    claude_cost_per = 0.08
    gemini_cost_per = 0.02

    kp_cost = n_papers * kimi_cost_per + claude_fb * claude_cost_per + gemini_fb * gemini_cost_per
    consensus_cost = n_papers * kimi_cost_per + n_papers * claude_cost_per + (n_papers // 4) * gemini_cost_per

    print(f"    Estimated cost: ~${kp_cost:.2f}")
    print(f"    Savings vs consensus: ~${consensus_cost - kp_cost:.2f} ({(1 - kp_cost/consensus_cost)*100:.0f}% reduction)")

    # ------------------------------------------------------------------
    # FINAL VERDICT
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("FINAL VERDICT")
    print(f"{'='*78}")

    print(f"""
  Symmetric Consensus:
    - N matched:    {consensus_stats['n']} obs from {consensus_stats['papers_matched']} papers
    - MAE:          {consensus_stats['mae']:.2f}%
    - Pearson r:    {consensus_stats['r']:.3f}
    - Direction:    {consensus_stats['direction_pct']:.1f}%
    - Overall eff:  {consensus_stats['overall_our']:+.2f}% (GT: {consensus_stats['overall_gt']:+.2f}%)

  Kimi-Primary + Fallback (obs-level):
    - N matched:    {kp_obs_stats['n']} obs from {kp_obs_stats['papers_matched']} papers
    - MAE:          {kp_obs_stats['mae']:.2f}%
    - Pearson r:    {kp_obs_stats['r']:.3f}
    - Direction:    {kp_obs_stats['direction_pct']:.1f}%
    - Overall eff:  {kp_obs_stats['overall_our']:+.2f}% (GT: {kp_obs_stats['overall_gt']:+.2f}%)

  Kimi Solo (no fallback):
    - N matched:    {kimi_stats['n']} obs from {kimi_stats['papers_matched']} papers
    - MAE:          {kimi_stats['mae']:.2f}%
    - Pearson r:    {kimi_stats['r']:.3f}
    - Direction:    {kimi_stats['direction_pct']:.1f}%
    - Overall eff:  {kimi_stats['overall_our']:+.2f}% (GT: {kimi_stats['overall_gt']:+.2f}%)
""")

    coverage_diff = consensus_stats['n'] - kp_obs_stats['n']
    mae_diff = consensus_stats['mae'] - kp_obs_stats['mae']

    if coverage_diff > 0 and mae_diff > 0:
        print(f"  -> Consensus provides +{coverage_diff} more matched observations")
        print(f"     but at +{mae_diff:.2f}pp higher MAE than Kimi-primary.")
        print(f"     This is a coverage-accuracy tradeoff.")
    elif coverage_diff > 0:
        print(f"  -> Consensus provides +{coverage_diff} more matched observations")
        print(f"     AND {abs(mae_diff):.2f}pp lower MAE. Consensus wins on both axes.")
    elif coverage_diff <= 0 and mae_diff >= 0:
        print(f"  -> Kimi-primary matches as many or more observations")
        print(f"     and has {abs(mae_diff):.2f}pp better MAE. Kimi-primary wins on both axes.")
    else:
        print(f"  -> Mixed result: Kimi-primary has {abs(coverage_diff)} more obs but {mae_diff:.2f}pp worse MAE.")

    print(f"\n  Key insight: The value of consensus is primarily in COVERAGE, not per-obs")
    print(f"  accuracy. On the observations that Kimi can extract, Kimi alone is often")
    print(f"  as good or better. But {n_fallback_calls} papers require fallback models to")
    print(f"  extract any usable data at all.")

    # ------------------------------------------------------------------
    # SAVE RESULTS
    # ------------------------------------------------------------------
    output_data = {
        'timestamp': datetime.now().isoformat(),
        'methods': {
            'consensus': {k: v for k, v in consensus_stats.items()},
            'kimi_primary_paper': {k: v for k, v in kp_stats.items()},
            'kimi_primary_obs': {k: v for k, v in kp_obs_stats.items()},
            'kimi_solo': {k: v for k, v in kimi_stats.items()},
            'claude_solo': {k: v for k, v in claude_stats.items()},
            'gemini_solo': {k: v for k, v in gemini_stats.items()},
        },
        'routing': {
            'paper_level': kp_routing,
            'obs_level': kp_obs_routing,
        },
        'paper_comparison': {
            'consensus_better': better_consensus,
            'kimi_primary_better': better_kp,
            'tied': tied,
        },
        'cost_estimate': {
            'consensus_cost': round(consensus_cost, 2),
            'kimi_primary_cost': round(kp_cost, 2),
            'savings': round(consensus_cost - kp_cost, 2),
        },
    }

    output_path = OUTPUT_DIR / "kimi_primary_simulation.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, default=str)
    print(f"\nResults saved to: {output_path}")


if __name__ == "__main__":
    main()
