"""
Validate Hui 2023 full 34-paper extraction against MOESM5 supplementary ground truth.

GT source: MOESM5_dataset.xlsx (Sheets 2-4: Soil, Foliar, Soil+Foliar application)
Extraction: output/hui2023_full_35/*_consensus.json
"""
import sys, json, math, csv, re
from pathlib import Path
from collections import defaultdict
from datetime import datetime

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
import numpy as np

GT_PATH = r"C:\Users\moshe\Dropbox\Testing metaanalyis program\Hui 2023 source data\supplementary\MOESM5_dataset.xlsx"
RESULTS_DIR = Path(r"C:\Users\moshe\Dropbox\Testing metaanalyis program\meta_analysis_extractor\output\hui2023_full_35")
OUTPUT_DIR = RESULTS_DIR

# Map paper filenames to GT study IDs PER SHEET.
# CRITICAL: Study IDs are sheet-specific (different numbering in each sheet).
# Using a single global ID pulled wrong-paper observations from other sheets.
# Format: {paper: {sheet_name: [study_ids]}}
PAPER_TO_SHEET_IDS = {
    "11_Zhao_2020": {
        "Data 2 Soil  application": [11],
        "Data 3 Foliar application": [18],
        "Data 4 Soil+Foliar application": [8],
    },
    "14_37_Liu_2019": {
        "Data 2 Soil  application": [14, 37],
    },
    "18_Zulfiqar_2020": {
        "Data 2 Soil  application": [18],
        "Data 3 Foliar application": [71],
    },
    "21_Wang_2012": {
        "Data 2 Soil  application": [21],
        "Data 3 Foliar application": [11],
    },
    "27_Zou_2012": {
        "Data 2 Soil  application": [27],
        "Data 3 Foliar application": [4],
        "Data 4 Soil+Foliar application": [11],
    },
    "3_Zhang_2012": {
        "Data 3 Foliar application": [3],  # 0 Grain Zn data in GT
    },
    "38_Yilmaz_1997": {
        "Data 2 Soil  application": [38],
        "Data 3 Foliar application": [64],
        "Data 4 Soil+Foliar application": [28],
    },
    "40_Bharti_2013": {
        "Data 2 Soil  application": [40],
        "Data 4 Soil+Foliar application": [14],
    },
    "42_Curtin_2008": {
        "Data 2 Soil  application": [42],  # 0 Grain Zn data in GT
    },
    "44_Cakmak_1997": {
        "Data 2 Soil  application": [44],
    },
    "45_Chattha_2017": {
        "Data 2 Soil  application": [45],
        "Data 3 Foliar application": [43],
        "Data 4 Soil+Foliar application": [15],
    },
    "46_Ghasal_2017": {
        "Data 2 Soil  application": [46],
        "Data 4 Soil+Foliar application": [17],
    },
    "49_Dawar_2022": {
        "Data 2 Soil  application": [49],  # 0 Grain Zn data in GT
    },
    "50_Erdal_2002": {
        "Data 2 Soil  application": [50],
    },
    "52_Forster_2018": {
        "Data 2 Soil  application": [52],
        "Data 3 Foliar application": [48],
    },
    "53_Grant_1998": {
        "Data 2 Soil  application": [53],  # 0 Grain Zn data in GT
    },
    "58_Kalayci_1999": {
        "Data 2 Soil  application": [58],
    },
    "59_Khoshgoftarmanesh_2013": {
        "Data 2 Soil  application": [59],
        "Data 3 Foliar application": [58],
    },
    "5_Yang_2011": {
        "Data 2 Soil  application": [5],
        "Data 3 Foliar application": [19],  # compound 19/33 in sheet
    },
    "61_Kumar_2018": {
        "Data 2 Soil  application": [61],  # 0 Grain Zn data
        "Data 3 Foliar application": [60],
        "Data 4 Soil+Foliar application": [13],
    },
    "62_Morshedi_2012": {
        "Data 2 Soil  application": [62],  # 0 Grain Zn data in GT
    },
    "63_Mosavian_2021": {
        "Data 2 Soil  application": [63],  # 0 Grain Zn data in GT
    },
    "65_Oliver_1994": {
        "Data 2 Soil  application": [65],  # 0 Grain Zn data in GT
    },
    "66_PahlavanRad_2009": {
        "Data 2 Soil  application": [66],
        "Data 3 Foliar application": [66],
    },
    "68_Peck_2008": {
        "Data 2 Soil  application": [68],
        "Data 4 Soil+Foliar application": [21],
    },
    "69_Ramzan_2020": {
        "Data 2 Soil  application": [69],
        "Data 3 Foliar application": [69],
    },
    "70_Rehman_2018": {
        "Data 2 Soil  application": [70],
    },
    "82_Torun_2001": {
        "Data 2 Soil  application": [82],  # 0 Grain Zn data in GT
    },
    "84_Yilmaz_1998": {
        "Data 2 Soil  application": [84],
    },
    "Dong_2018": {
        "Data 3 Foliar application": [8, 31],
    },
    "Li_2013": {
        "Data 2 Soil  application": [19],
        "Data 3 Foliar application": [10],
        "Data 4 Soil+Foliar application": [3],
    },
    "Liu_2014": {
        "Data 2 Soil  application": [16],
        "Data 3 Foliar application": [1],
        "Data 4 Soil+Foliar application": [2],
    },
    "Rashid_2019": {
        "Data 2 Soil  application": [72],
    },
    "Zhang_2017": {
        "Data 2 Soil  application": [35],
    },
}

# Flat view for iteration (paper list)
PAPER_IDS = list(PAPER_TO_SHEET_IDS.keys())

# Sheet configs: column indices for Grain Zn
SHEET_COLS = {
    "Data 2 Soil  application": {"study_id": 1, "pub": 2, "n": 21, "zn_ctrl": 33, "zn_treat": 34, "zn_effect": 35},
    "Data 3 Foliar application": {"study_id": 1, "pub": 2, "n": 21, "zn_ctrl": 33, "zn_treat": 34, "zn_effect": 35},
    "Data 4 Soil+Foliar application": {"study_id": 1, "pub": 2, "n": 4, "zn_ctrl": 14, "zn_treat": 15, "zn_effect": 16},
}


def load_gt():
    """Load GT from MOESM5, returning {(sheet_name, study_id): [obs_list]}.

    CRITICAL: Study IDs are sheet-specific. The same ID number refers to
    different papers in different sheets. We key by (sheet, study_id) tuple.
    """
    wb = openpyxl.load_workbook(GT_PATH, data_only=True)
    gt_by_sheet_study = defaultdict(list)

    for sname, cols in SHEET_COLS.items():
        ws = wb[sname]
        app_type = sname.split()[1]  # "2"=Soil, "3"=Foliar, "4"=Soil+Foliar

        row2 = [str(c.value)[:50] if c.value else '' for c in ws[2]]

        # Find Grain Zn columns dynamically
        zn_ctrl_col = None
        for i, h in enumerate(row2):
            if 'Grain Zn concentration' in h:
                zn_ctrl_col = i
                break

        if zn_ctrl_col is None:
            print(f"  WARNING: Could not find Grain Zn column in {sname}")
            continue

        zn_treat_col = zn_ctrl_col + 1
        zn_effect_col = zn_ctrl_col + 2

        # Find n column
        n_col = None
        for i, h in enumerate(row2):
            if 'replicates' in h.lower() or 'number' in h.lower():
                n_col = i
                break

        seen_rows = set()  # Deduplicate compound IDs (e.g., "19/33")
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
            study_id = row[1]
            pub = str(row[2]).strip() if row[2] else ""

            if not pub or study_id is None:
                continue

            try:
                ctrl = float(row[zn_ctrl_col]) if row[zn_ctrl_col] is not None else None
                treat = float(row[zn_treat_col]) if row[zn_treat_col] is not None else None
                effect = float(row[zn_effect_col]) if row[zn_effect_col] is not None else None
                n = int(row[n_col]) if n_col and row[n_col] is not None else None
            except (ValueError, TypeError):
                continue

            if ctrl is None or treat is None or ctrl <= 0:
                continue

            # Handle compound study IDs like "19/33"
            sid_str = str(study_id)
            study_ids_parsed = []
            if '/' in sid_str:
                for part in sid_str.split('/'):
                    try:
                        study_ids_parsed.append(int(part.strip()))
                    except ValueError:
                        pass
            else:
                try:
                    study_ids_parsed.append(int(float(sid_str)))
                except ValueError:
                    continue

            obs = {
                'study_id': study_ids_parsed[0] if study_ids_parsed else 0,
                'all_study_ids': study_ids_parsed,
                'publication': pub[:100],
                'ctrl': ctrl,
                'treat': treat,
                'effect': effect,
                'n': n,
                'sheet': sname,
                'app_type': app_type,
                'row_key': f"{sname}_{row_idx}",  # Unique row key for dedup
            }

            # Add to each parsed study_id under this sheet
            for sid_val in study_ids_parsed:
                gt_by_sheet_study[(sname, sid_val)].append(obs)

    return gt_by_sheet_study


def get_gt_for_paper(paper_id, gt_by_sheet_study):
    """Get deduplicated GT observations for a paper using per-sheet mapping."""
    sheet_ids = PAPER_TO_SHEET_IDS.get(paper_id, {})
    seen_rows = set()
    gt_rows = []

    for sheet_name, study_ids in sheet_ids.items():
        for sid in study_ids:
            for obs in gt_by_sheet_study.get((sheet_name, sid), []):
                row_key = obs['row_key']
                if row_key not in seen_rows:
                    seen_rows.add(row_key)
                    gt_rows.append(obs)

    return gt_rows


def load_extraction(paper_id):
    """Load consensus observations for a paper."""
    path = RESULTS_DIR / f"{paper_id}_consensus.json"
    if not path.exists():
        return []

    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    obs_list = data.get('consensus_observations', [])

    # Filter to Zn grain observations
    zn_obs = []
    for obs in obs_list:
        el = (obs.get('element') or '').lower()
        tissue = (obs.get('tissue') or '').lower()
        if 'zn' not in el and 'zinc' not in el:
            continue

        ctrl = obs.get('control_mean')
        treat = obs.get('treatment_mean')
        if ctrl is None or treat is None:
            continue

        try:
            ctrl = float(ctrl)
            treat = float(treat)
        except (ValueError, TypeError):
            continue

        if ctrl <= 0:
            continue

        zn_obs.append({
            'ctrl': ctrl,
            'treat': treat,
            'effect_pct': (treat - ctrl) / ctrl * 100,
            'element': obs.get('element', 'Zn'),
            'tissue': obs.get('tissue', 'grain'),
            'source': obs.get('data_source', ''),
            'treatment_desc': obs.get('treatment_description', ''),
        })

    return zn_obs


def match_observations(extracted, gt_rows, tolerance=0.15):
    """Match extracted obs to GT rows by control/treatment values.
    Returns list of (extracted, gt, error) tuples.
    """
    matches = []
    used_gt = set()
    used_ext = set()

    # Sort by closest match first
    candidates = []
    for i, ext in enumerate(extracted):
        for j, gt in enumerate(gt_rows):
            ctrl_err = abs(ext['ctrl'] - gt['ctrl']) / max(gt['ctrl'], 0.01)
            treat_err = abs(ext['treat'] - gt['treat']) / max(gt['treat'], 0.01)
            combined = ctrl_err + treat_err
            candidates.append((combined, i, j, ctrl_err, treat_err))

    candidates.sort()

    for combined, i, j, ctrl_err, treat_err in candidates:
        if i in used_ext or j in used_gt:
            continue
        if combined <= tolerance * 2:  # Both within tolerance
            ext = extracted[i]
            gt = gt_rows[j]
            gt_effect = (gt['treat'] - gt['ctrl']) / gt['ctrl'] * 100
            ext_effect = ext['effect_pct']
            abs_err = abs(ext_effect - gt_effect)
            matches.append({
                'ext_ctrl': ext['ctrl'],
                'ext_treat': ext['treat'],
                'gt_ctrl': gt['ctrl'],
                'gt_treat': gt['treat'],
                'ext_effect': ext_effect,
                'gt_effect': gt_effect,
                'abs_error': abs_err,
                'ctrl_err': ctrl_err,
                'treat_err': treat_err,
                'tissue': ext.get('tissue', 'grain'),
                'app_type': gt.get('app_type', ''),
            })
            used_gt.add(j)
            used_ext.add(i)

    return matches


def compute_metrics(matches):
    """Compute validation metrics from matched observations."""
    if not matches:
        return {}

    ext_effects = [m['ext_effect'] for m in matches]
    gt_effects = [m['gt_effect'] for m in matches]
    errors = [m['abs_error'] for m in matches]

    # Pearson correlation
    ext_arr = np.array(ext_effects)
    gt_arr = np.array(gt_effects)
    if len(ext_arr) > 1 and np.std(ext_arr) > 0 and np.std(gt_arr) > 0:
        r = np.corrcoef(ext_arr, gt_arr)[0, 1]
    else:
        r = None

    # Direction accuracy
    n_correct_dir = sum(1 for e, g in zip(ext_effects, gt_effects)
                        if (e > 0 and g > 0) or (e < 0 and g < 0) or (abs(e) < 1 and abs(g) < 1))

    mae = np.mean(errors)
    median_ae = np.median(errors)
    within_5 = sum(1 for e in errors if e <= 5)
    within_10 = sum(1 for e in errors if e <= 10)
    within_20 = sum(1 for e in errors if e <= 20)

    return {
        'n': len(matches),
        'pearson_r': round(r, 3) if r is not None else None,
        'mae_pct': round(float(mae), 2),
        'median_ae_pct': round(float(median_ae), 2),
        'direction_accuracy': round(n_correct_dir / len(matches), 3),
        'within_5pp': round(within_5 / len(matches), 3),
        'within_10pp': round(within_10 / len(matches), 3),
        'within_20pp': round(within_20 / len(matches), 3),
    }


def main():
    print("=" * 70)
    print("HUI 2023 FULL VALIDATION (34 papers)")
    print("=" * 70)

    # Load GT
    print("\n[1] Loading ground truth from MOESM5...")
    gt_by_sheet_study = load_gt()
    # Count unique observations (deduplicate compound IDs)
    all_row_keys = set()
    for obs_list in gt_by_sheet_study.values():
        for obs in obs_list:
            all_row_keys.add(obs['row_key'])
    total_gt = len(all_row_keys)
    print(f"  {len(gt_by_sheet_study)} (sheet, study_id) entries, {total_gt} unique observations")

    # Process each paper
    print("\n[2] Matching extractions to GT...")
    all_matches = []
    paper_stats = []

    for paper_id in sorted(PAPER_IDS):
        extracted = load_extraction(paper_id)

        # Get GT using per-sheet study ID mapping (avoids cross-sheet collisions)
        gt_rows = get_gt_for_paper(paper_id, gt_by_sheet_study)

        if not gt_rows:
            paper_stats.append({
                'paper': paper_id,
                'status': 'no_gt',
                'n_extracted': len(extracted),
                'n_gt': 0,
                'n_matched': 0,
            })
            continue

        if not extracted:
            paper_stats.append({
                'paper': paper_id,
                'status': 'no_extraction',
                'n_extracted': 0,
                'n_gt': len(gt_rows),
                'n_matched': 0,
            })
            continue

        matches = match_observations(extracted, gt_rows)
        for m in matches:
            m['paper_id'] = paper_id  # tag each match with its source paper
        all_matches.extend(matches)

        metrics = compute_metrics(matches)
        paper_stats.append({
            'paper': paper_id,
            'status': 'matched',
            'n_extracted': len(extracted),
            'n_gt': len(gt_rows),
            'n_matched': len(matches),
            'capture_rate': round(len(matches) / len(gt_rows) * 100, 1) if gt_rows else 0,
            **metrics,
        })

    # Overall metrics
    overall = compute_metrics(all_matches)

    # Overall effect comparison
    gt_effects = [m['gt_effect'] for m in all_matches]
    ext_effects = [m['ext_effect'] for m in all_matches]
    gt_mean = np.mean(gt_effects) if gt_effects else 0
    ext_mean = np.mean(ext_effects) if ext_effects else 0

    # Print results
    print(f"\n{'=' * 70}")
    print("RESULTS")
    print(f"{'=' * 70}")
    print(f"\n  Papers processed: {len(PAPER_TO_SHEET_IDS)}")
    print(f"  Papers with matches: {sum(1 for p in paper_stats if p['n_matched'] > 0)}")
    print(f"  Total GT observations: {sum(p['n_gt'] for p in paper_stats)}")
    print(f"  Total extracted Zn obs: {sum(p['n_extracted'] for p in paper_stats)}")
    print(f"  Total matched: {len(all_matches)}")

    print(f"\n  OVERALL METRICS:")
    print(f"    Pearson r:          {overall.get('pearson_r', 'N/A')}")
    print(f"    MAE:                {overall.get('mae_pct', 'N/A')}%")
    print(f"    Median AE:          {overall.get('median_ae_pct', 'N/A')}%")
    print(f"    Direction accuracy:  {overall.get('direction_accuracy', 'N/A')}")
    print(f"    Within 5pp:         {overall.get('within_5pp', 'N/A')}")
    print(f"    Within 10pp:        {overall.get('within_10pp', 'N/A')}")
    print(f"    Within 20pp:        {overall.get('within_20pp', 'N/A')}")

    print(f"\n  OVERALL EFFECT:")
    print(f"    GT mean effect:     {gt_mean:.2f}%")
    print(f"    Extracted mean:     {ext_mean:.2f}%")
    print(f"    Difference:         {abs(gt_mean - ext_mean):.2f}pp")

    # Per-paper table
    print(f"\n  PER-PAPER RESULTS:")
    print(f"  {'Paper':<35} {'GT':>4} {'Ext':>4} {'Match':>5} {'Cap%':>5} {'r':>6} {'MAE':>6} {'Dir':>5}")
    print(f"  {'-'*75}")
    for p in sorted(paper_stats, key=lambda x: -x.get('n_matched', 0)):
        r_str = f"{p.get('pearson_r', '')}" if p.get('pearson_r') is not None else "-"
        mae_str = f"{p.get('mae_pct', '')}" if p.get('mae_pct') is not None else "-"
        dir_str = f"{p.get('direction_accuracy', '')}" if p.get('direction_accuracy') is not None else "-"
        cap_str = f"{p.get('capture_rate', 0)}" if p.get('n_gt', 0) > 0 else "-"
        print(f"  {p['paper']:<35} {p.get('n_gt',0):>4} {p.get('n_extracted',0):>4} {p.get('n_matched',0):>5} {cap_str:>5} {r_str:>6} {mae_str:>6} {dir_str:>5}")

    # Save results
    report = {
        'timestamp': datetime.now().isoformat(),
        'dataset': 'Hui et al. 2023 - Zinc/Wheat (34 papers)',
        'papers_processed': len(PAPER_TO_SHEET_IDS),
        'papers_matched': sum(1 for p in paper_stats if p['n_matched'] > 0),
        'total_gt': sum(p['n_gt'] for p in paper_stats),
        'total_extracted_zn': sum(p['n_extracted'] for p in paper_stats),
        'total_matched': len(all_matches),
        'metrics': overall,
        'overall_effect': {
            'gt_mean_pct': round(float(gt_mean), 2),
            'extracted_mean_pct': round(float(ext_mean), 2),
            'difference_pp': round(abs(float(gt_mean - ext_mean)), 2),
        },
        'paper_stats': paper_stats,
    }

    report_path = OUTPUT_DIR / 'validation_report.json'
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n  Report saved to: {report_path}")

    # Save matches CSV
    csv_path = OUTPUT_DIR / 'validation_matches.csv'
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['paper_id', 'ext_ctrl', 'ext_treat', 'gt_ctrl', 'gt_treat',
                         'ext_effect', 'gt_effect', 'abs_error', 'tissue', 'app_type'])
        for m in all_matches:
            writer.writerow([
                m.get('paper_id', 'unknown'),
                m['ext_ctrl'], m['ext_treat'], m['gt_ctrl'], m['gt_treat'],
                round(m['ext_effect'], 2), round(m['gt_effect'], 2),
                round(m['abs_error'], 2), m['tissue'], m['app_type'],
            ])
    print(f"  Matches saved to: {csv_path}")


if __name__ == '__main__':
    main()
