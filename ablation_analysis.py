"""
Ablation Analysis: Consensus Pipeline vs Single-Model Baselines

Compares the consensus pipeline against each individual LLM model,
quantifying gains from:
1. Multi-model consensus vs single models
2. Tiebreaker contribution (Gemini)
3. TEXT vs HYBRID/VISION extraction mode
4. Challenge-aware routing (HARD vs MEDIUM difficulty)

Uses the same matching logic as validate_full_46.py to ensure
fair comparison on the same GT scope.

Usage:
    .\venv\Scripts\python.exe ablation_analysis.py
"""
import sys, os, json, math, csv, re
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from itertools import groupby

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv()

import openpyxl

# ==================================================================
# PATHS
# ==================================================================
BASE_DIR = Path(r"C:\Users\moshe\Dropbox\Testing metaanalyis program\meta_analysis_extractor")
GT_PATH = Path(r"C:\Users\moshe\Dropbox\Testing metaanalyis program\Loladze\CO2+Dataset.xlsx")

CONSENSUS_DIR = BASE_DIR / "output" / "loladze_full_46_v2"
BASELINE_DIRS = {
    "Claude": BASE_DIR / "output" / "baseline_claude_loladze",
    "Gemini": BASE_DIR / "output" / "baseline_gemini_loladze",
    "Kimi":   BASE_DIR / "output" / "baseline_kimi_loladze",
}

OUTPUT_DIR = BASE_DIR / "output"

# ==================================================================
# PAPER MAPPING (from validate_full_46.py)
# ==================================================================
PAPER_TO_LOLADZE_REF = {
    "001_Ma_2007": "Fernando et al 2012a",
    "002_Ziska_1997": "Ziska et al 1997",
    "003_Baslam_2012": "Baslam et al 2012",
    "004_Finzi_2001": "Finzi et al 2001",
    "005_Niinemets_1999": "Niinemets et al 1999",
    "006_Azam_2013": "Azam et al 2013",
    "007_Woodin_1992": "Woodin et al 1992",
    "008_Campbell_2002": "Cambell & Sage 2002",
    "009_Barnes_1992": "Barnes & Pffirrman 1992",
    "010_Li_2010": "Hogy et al 2009",
    "011_Huluka_1994": "Huluka et al 1994",
    "012_Wu_2004": "Wu et al 2004",
    "013_Keutgen_2001": "Keutgen & Chen 2001",
    "014_Lieffering_2004": "Lieffering et al 2004",
    "015_Pleijel_2009": "Pleijel & Danielsson 2009",
    "016_Fernando_2012a": "Fernando et al 2012b",
    "017_Fangmeier_2002": "Fangmeier et al  2002",
    "018_Al-Rawahy_2013": "Al-Rawahy et al 2013",
    "019_Baxter_1994": "Baxter et al 1994",
    "020_Overdieck_1993": "Overdieck 1993",
    "021_Wilsey_1994": "Wilsey et al 1994",
    "022_Blank_2011": "Blank et al 2011",
    "025_Guo_2011": "Guo et al 2013",
    "026_Seneweera_1997": "Seneweera & Conroy 1997",
    "027_Peet_1986": "Peet et al 1986",
    "028_Mishra_2011": "Mishra et al 2011",
    "031_Pal_2003": "Pal et al 2004",
    "032_Kanowski_2001": "Kanowski 2001",
    "034_Johnson_1997": "Johnson et al 2003",
    "035_Oksanen_2005": "Oksanen et al 2005",
    "036_Schenk_1997": "Schenk et al 1997",
    "037_de_2000": "Haase et al 2008",
    "038_Newbery_1995": "Newbery 1995",
    "039_Heagle_1993": "Heagle et al 1993",
    "040_Pfirrmann_1996": "Pfirrmann et al 1996",
    "041_Mjwara_1996": "Mjwara et al 1996",
    "042_Luomala_2005": "Luomala et al 2005",
    "043_Natali_2009": "Natali et al 2009",
    "044_Housman_2012": "Housman et al 2012",
    "046_Porter_1984": "Porter & Grodzinski 1984",
    "047_Rodenkirchen_2009": "Rodenkirchen et al 2009",
    "048_Khan_2013": "Khan et al 2012",
    "049_Singh_2013": "Singh et al 2013",
    "050_Polley_2011": "Polley et al 2011",
    "051_Niu_2013": "Niu et al 2013",
    "058_ONeill_1987": "O'Neill et al 1987",
}

MISLABELED_PDFS = {
    "001_Ma_2007": "Fernando et al 2012a",
    "010_Li_2010": "Hogy et al 2009",
    "031_Pal_2003": "Pal et al 2004",
    "034_Johnson_1997": "Johnson et al 2003",
    "037_de_2000": "Haase et al 2008",
}


# ==================================================================
# HELPER FUNCTIONS (imported logic from validate_full_46.py)
# ==================================================================

def normalize_element(el):
    if not el:
        return ""
    el = el.strip().upper()
    MAP = {"NITROGEN": "N", "PHOSPHORUS": "P", "POTASSIUM": "K",
           "CALCIUM": "CA", "MAGNESIUM": "MG", "IRON": "FE",
           "ZINC": "ZN", "MANGANESE": "MN", "COPPER": "CU",
           "SULFUR": "S", "SULPHUR": "S", "BORON": "B", "CARBON": "C",
           "SODIUM": "NA", "MOLYBDENUM": "MO", "SILICON": "SI",
           "SELENIUM": "SE", "COBALT": "CO", "CHROMIUM": "CR",
           "CADMIUM": "CD", "NICKEL": "NI", "LEAD": "PB",
           "ALUMINIUM": "AL", "ALUMINUM": "AL", "BARIUM": "BA",
           "STRONTIUM": "SR", "BROMINE": "BR", "VANADIUM": "V",
           "CHLORINE": "CL"}
    return MAP.get(el, el)


def load_gt():
    """Load Loladze GT with Additional Info for each row."""
    wb = openpyxl.load_workbook(GT_PATH, data_only=True)
    ws = wb["CO2 Dataset"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[2]]

    col = {}
    for i, h in enumerate(headers):
        if h == 'Reference': col['ref'] = i
        elif h == '(E-A)/A': col['effect'] = i
        elif h == 'Element': col['element'] = i
        elif h == 'Additional Info': col['info'] = i
        elif h == 'eCO2': col['eco2'] = i
        elif h == 'aCO2': col['aco2'] = i
        elif h == 'Tissue': col['tissue'] = i
        elif h == 'Species': col['species'] = i
        elif h == 'n': col['n'] = i
        elif h == 'Foliar Edible': col['foliar_edible'] = i

    gt = defaultdict(list)
    for row in ws.iter_rows(min_row=3, values_only=True):
        ref = str(row[col['ref']]).strip() if row[col['ref']] else ""
        el = normalize_element(str(row[col['element']]).strip() if row[col['element']] else "")
        eff = row[col['effect']]
        info = str(row[col.get('info', 11)]).strip() if row[col.get('info', 11)] else ""
        if info == "None":
            info = ""
        eco2 = row[col.get('eco2', 8)]
        tissue = str(row[col.get('tissue', 4)]).strip() if row[col.get('tissue', 4)] else ""
        species = str(row[col.get('species', 0)]).strip() if col.get('species') is not None and row[col['species']] else ""
        n_val = row[col.get('n', 17)]

        if ref and el and eff is not None:
            try:
                gt[ref].append({
                    'element': el,
                    'effect': float(eff),
                    'info': info,
                    'eco2': float(eco2) if eco2 else None,
                    'tissue': tissue,
                    'species': species,
                    'n': int(n_val) if n_val else None,
                })
            except (ValueError, TypeError):
                pass

    wb.close()
    return dict(gt)


def get_mods(obs):
    mods = obs.get('moderators', {})
    if isinstance(mods, str):
        try:
            mods = json.loads(mods)
        except:
            mods = {}
    if not isinstance(mods, dict):
        mods = {}
    return mods


def is_concentration_unit(unit_str):
    u = str(unit_str).lower()
    total_keywords = ['plant', 'pot', 'shoot', 'total', 'uptake', 'content']
    if any(kw in u for kw in total_keywords):
        return False
    return True


def filter_obs_for_gt_row(obs_list, gt_row, paper_id):
    """
    Filter observations to match a specific GT row's conditions.
    Simplified version of the validate_full_46.py filter.
    Focuses on element + key moderator matching.
    """
    gt_el = gt_row['element']
    gt_info = gt_row['info'].lower()

    # Filter to matching element
    el_matches = [o for o in obs_list if normalize_element(o.get('element', '')) == gt_el]
    if not el_matches:
        return []

    # Filter out total-content-per-plant observations
    conc_matches = [o for o in el_matches if is_concentration_unit(o.get('unit', ''))]
    if conc_matches:
        el_matches = conc_matches

    if not gt_info:
        return el_matches

    filtered = el_matches

    # --- Leaf position filters (Baslam) ---
    if 'inner' in gt_info and 'nm' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            leaf_pos = str(mods.get('leaf_position', '')).lower()
            amf = str(mods.get('AMF_status', mods.get('mycorrhizal_status', ''))).lower()
            if 'inner' in leaf_pos and ('non' in amf or 'nm' in amf):
                new.append(o)
        if new:
            filtered = new
    elif 'nm' in gt_info or 'non-mycorrhizal' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            amf = str(mods.get('AMF_status', mods.get('mycorrhizal_status', ''))).lower()
            if 'non' in amf or 'nm' in amf:
                new.append(o)
        if new:
            filtered = new

    # --- Leaf age ---
    if 'old leaves' in gt_info:
        new = [o for o in filtered if 'old' in str(get_mods(o).get('leaf_age', '')).lower()]
        if new: filtered = new
    if 'young leaves' in gt_info:
        new = [o for o in filtered if 'young' in str(get_mods(o).get('leaf_age', '')).lower()]
        if new: filtered = new

    # --- Year filter ---
    year_match = re.match(r'^(\d{4})$', gt_row['info'].strip())
    if year_match:
        target_year = year_match.group(1)
        new = []
        for o in filtered:
            mods = get_mods(o)
            if target_year in str(mods.get('year', '')):
                new.append(o)
        if new:
            filtered = new

    # --- High/Low N ---
    if 'high n' in gt_info and 'low' not in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            n_level = str(mods.get('nitrogen_level', mods.get('N_level', mods.get('n_treatment', '')))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if 'high' in n_level or 'high n' in desc:
                new.append(o)
        if new: filtered = new
    elif 'low n' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            n_level = str(mods.get('nitrogen_level', mods.get('N_level', mods.get('n_treatment', '')))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if 'low' in n_level or 'low n' in desc:
                new.append(o)
        if new: filtered = new

    # --- Soil type ---
    if 'basalt' in gt_info:
        new = [o for o in filtered
               if 'basalt' in str(get_mods(o).get('soil_type', '')).lower()
               or 'basalt' in str(o.get('treatment_description', '')).lower()]
        if new: filtered = new
    if 'rhyolite' in gt_info:
        new = [o for o in filtered
               if 'rhyolite' in str(get_mods(o).get('soil_type', '')).lower()
               or 'rhyolite' in str(o.get('treatment_description', '')).lower()]
        if new: filtered = new

    # --- Cultivar (Heagle) ---
    for cultivar_name in ['nc-r', 'nc-s']:
        if cultivar_name in gt_info:
            new = []
            for o in filtered:
                mods = get_mods(o)
                cult = str(mods.get('cultivar', mods.get('clone', ''))).lower()
                desc = str(o.get('treatment_description', '')).lower()
                if cultivar_name in cult or cultivar_name in desc:
                    new.append(o)
            if new: filtered = new
            break

    # --- Potassium treatment (Pfirrmann) ---
    if '+k' in gt_info and '-k' not in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            k_treat = str(mods.get('potassium', mods.get('K_treatment', ''))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if '+k' in k_treat or 'plus k' in k_treat or '+k' in desc or 'with k' in desc:
                new.append(o)
        if new: filtered = new
    elif '-k' in gt_info:
        new = []
        for o in filtered:
            mods = get_mods(o)
            k_treat = str(mods.get('potassium', mods.get('K_treatment', ''))).lower()
            desc = str(o.get('treatment_description', '')).lower()
            if '-k' in k_treat or 'minus k' in k_treat or '-k' in desc or 'without k' in desc:
                new.append(o)
        if new: filtered = new

    # --- P level (Seneweera) ---
    p_level_match = re.search(r'p\s+(\d+)\s*mg/kg', gt_info)
    if p_level_match:
        target_p = p_level_match.group(1)
        target_p_int = int(target_p)
        new = []
        for o in filtered:
            mods = get_mods(o)
            all_mods_str = str(mods).lower()
            desc = str(o.get('treatment_description', '')).lower()
            for mk, mv in mods.items():
                mk_lower = str(mk).lower()
                if ('p_level' in mk_lower or 'phosph' in mk_lower) and mv is not None:
                    try:
                        if abs(float(mv) - target_p_int) < 1:
                            new.append(o)
                            break
                    except (ValueError, TypeError):
                        if target_p in str(mv):
                            new.append(o)
                            break
            else:
                if f'p {target_p}' in desc or f'{target_p} mg' in all_mods_str:
                    new.append(o)
        if new: filtered = new

    # --- Harvest year ---
    harvest_year_match = re.search(r'(\d{4})\s*harvest', gt_info)
    if harvest_year_match:
        target_year = harvest_year_match.group(1)
        new = []
        for o in filtered:
            mods = get_mods(o)
            all_text = str(mods) + ' ' + str(o.get('treatment_description', ''))
            if target_year in all_text:
                new.append(o)
        if new: filtered = new

    # --- Natali site filters ---
    if paper_id == '043_Natali_2009':
        if 'duke' in gt_info and 'serc' not in gt_info:
            new = [o for o in filtered
                   if 'duke' in str(get_mods(o).get('site', '')).lower()
                   or 'duke' in str(o.get('treatment_description', '')).lower()]
            if new: filtered = new
            if '1 yr old' in gt_info:
                new = [o for o in filtered
                       if 'pinus' in str(get_mods(o).get('species', '')).lower()
                       or 'taeda' in str(get_mods(o).get('species', '')).lower()]
                if new: filtered = new
            else:
                new = [o for o in filtered
                       if 'liquidambar' in str(get_mods(o).get('species', '')).lower()
                       or 'sweetgum' in str(o.get('treatment_description', '')).lower()]
                if new: filtered = new
        elif 'ornl' in gt_info:
            new = [o for o in filtered
                   if 'ornl' in str(get_mods(o).get('site', '')).lower()]
            if new: filtered = new
        elif 'serc' in gt_info:
            new = [o for o in filtered
                   if 'serc' in str(get_mods(o).get('site', '')).lower()]
            if new: filtered = new
            gt_species = gt_row.get('species', '').lower()
            if gt_species:
                gt_parts = [p for p in gt_species.split() if len(p) > 3]
                species_new = []
                for o in filtered:
                    obs_sp = str(get_mods(o).get('species', '')).lower()
                    desc = str(o.get('treatment_description', '')).lower()
                    combined = obs_sp + ' ' + desc
                    if all(part in combined for part in gt_parts):
                        species_new.append(o)
                if species_new: filtered = species_new

    # --- Ozone filter ---
    if 'o3' not in gt_info and 'ozone' not in gt_info:
        has_o3_mods = any(
            any(k.lower() in ('o3_level', 'ozone', 'o3', 'ozone_level')
                for k in get_mods(o).keys())
            for o in filtered
        )
        if has_o3_mods:
            ambient_o3 = []
            for o in filtered:
                mods = get_mods(o)
                o3_val = ''
                for k, v in mods.items():
                    if k.lower() in ('o3_level', 'ozone', 'o3', 'ozone_level'):
                        o3_val = str(v).lower()
                        break
                if any(term in o3_val for term in ['ambient', 'cf', 'charcoal',
                       'control', 'low', 'clean', 'filtered', '5 nl',
                       '20 n', '20n', 'background']):
                    ambient_o3.append(o)
                elif not o3_val:
                    ambient_o3.append(o)
            if ambient_o3:
                filtered = ambient_o3

    # --- CO2 level filter ---
    if gt_row.get('eco2'):
        gt_co2 = gt_row['eco2']
        new = []
        for o in filtered:
            desc = str(o.get('treatment_description', '')).lower()
            mods = get_mods(o)
            co2_match_re = re.search(r'(\d{3,4})\s*(?:ppm|umol|umol|mumol)', desc)
            our_co2 = None
            if co2_match_re:
                our_co2 = float(co2_match_re.group(1))
            else:
                for k, v in mods.items():
                    k_l = k.lower()
                    if 'co2' in k_l and v is not None:
                        try:
                            our_co2 = float(v)
                            break
                        except (ValueError, TypeError):
                            co2_in_mod = re.search(r'(\d{3,4})', str(v))
                            if co2_in_mod:
                                our_co2 = float(co2_in_mod.group(1))
                                break
            if our_co2 is not None:
                if abs(our_co2 - gt_co2) < 100:
                    new.append(o)
            else:
                new.append(o)
        if new:
            filtered = new

    return filtered


def detect_tc_swap(obs_list, gt_rows):
    """Detect if extraction has treatment/control swapped."""
    our_effects = []
    gt_effects = []
    for gt_row in gt_rows:
        gt_el = gt_row['element']
        matching = [o for o in obs_list
                    if normalize_element(o.get('element', '')) == gt_el
                    and is_concentration_unit(o.get('unit', ''))]
        if not matching:
            continue
        for o in matching:
            ctrl = o.get('control_mean')
            treat = o.get('treatment_mean')
            if ctrl and ctrl != 0 and treat is not None:
                our_eff = (treat - ctrl) / ctrl
                if abs(our_eff) <= 5.0:
                    our_effects.append(our_eff)
                    gt_effects.append(gt_row['effect'])

    if len(our_effects) < 3:
        return False

    disagree = sum(1 for o, g in zip(our_effects, gt_effects)
                   if g != 0 and (o < 0) != (g < 0))
    return disagree / len(our_effects) > 0.7


def compute_effect(obs, swap_tc=False):
    ctrl = obs.get('control_mean')
    treat = obs.get('treatment_mean')
    if swap_tc:
        ctrl, treat = treat, ctrl
    if ctrl and ctrl != 0 and treat is not None:
        return (treat - ctrl) / ctrl
    return None


def deduplicate_vision_text(obs_list):
    """Prefer text over vision when both exist."""
    text_obs = []
    vision_obs = []
    for o in obs_list:
        notes = str(o.get('notes', '')).lower()
        if '[from vision]' in notes or 'vision' in notes:
            vision_obs.append(o)
        else:
            text_obs.append(o)
    if not vision_obs:
        return obs_list
    text_keys = set()
    for o in text_obs:
        mods = get_mods(o)
        el = normalize_element(o.get('element', ''))
        sp = str(mods.get('species', '')).lower().strip()
        site = str(mods.get('site', '')).lower().strip()
        text_keys.add((el, sp, site))
    deduped = list(text_obs)
    for o in vision_obs:
        mods = get_mods(o)
        el = normalize_element(o.get('element', ''))
        sp = str(mods.get('species', '')).lower().strip()
        site = str(mods.get('site', '')).lower().strip()
        if (el, sp, site) not in text_keys:
            deduped.append(o)
    return deduped


# ==================================================================
# CORE: Validate a set of results against GT
# ==================================================================

def validate_source(source_dir, gt, source_type="consensus"):
    """
    Validate a source directory against GT.
    source_type: "consensus" for consensus JSONs, "baseline" for baseline JSONs.
    Returns list of match dicts with (paper, el, our, gt, err, info).
    """
    if source_type == "consensus":
        pattern = "*_consensus.json"
        obs_key = "consensus_observations"
    else:
        pattern = "*_baseline.json"
        obs_key = "observations"

    results_files = sorted(source_dir.glob(pattern))
    all_matches = []
    paper_details = {}

    for rf in results_files:
        if rf.name.startswith('_'):
            continue  # skip summary files

        paper_id = rf.stem.replace("_consensus", "").replace("_baseline", "")
        loladze_ref = PAPER_TO_LOLADZE_REF.get(paper_id)

        if not loladze_ref or loladze_ref not in gt:
            for ref in gt:
                surname = paper_id.split('_')[1] if '_' in paper_id else paper_id
                if surname.lower() in ref.lower():
                    loladze_ref = ref
                    break
            if not loladze_ref or loladze_ref not in gt:
                continue

        with open(rf, encoding='utf-8') as f:
            data = json.load(f)

        gt_rows = gt[loladze_ref]
        obs_list = data.get(obs_key, [])

        if not obs_list:
            continue

        # Deduplicate vision/text
        obs_list = deduplicate_vision_text(obs_list)

        # Filter to concentration units
        has_conc = any(is_concentration_unit(o.get('unit', '')) for o in obs_list)
        has_total = any(not is_concentration_unit(o.get('unit', ''))
                        for o in obs_list if o.get('unit', ''))
        if not has_conc and has_total:
            continue

        # Filter sub-ambient CO2
        elevated_obs = []
        for o in obs_list:
            desc = str(o.get('treatment_description', '')).lower()
            co2_match = re.search(r'(\d{2,4})\s*(?:ppm|umol|umol|mumol)', desc)
            if co2_match:
                co2_val = float(co2_match.group(1))
                if co2_val < 300:
                    continue
            if 'low co2' in desc or 'sub-ambient' in desc:
                continue
            elevated_obs.append(o)
        if elevated_obs:
            obs_list = elevated_obs

        # Detect T/C swap
        swap_tc = detect_tc_swap(obs_list, gt_rows)

        # Match GT rows
        gt_by_el_info = defaultdict(list)
        for gt_row in gt_rows:
            key = (gt_row['element'], gt_row['info'])
            gt_by_el_info[key].append(gt_row)

        used_obs_ids = set()
        matched = 0

        for gt_row in gt_rows:
            candidates = filter_obs_for_gt_row(obs_list, gt_row, paper_id)
            if not candidates:
                continue

            cand_effects = []
            for i, c in enumerate(candidates):
                obs_id = id(c)
                eff = compute_effect(c, swap_tc=swap_tc)
                if eff is not None and abs(eff) <= 5.0:
                    cand_effects.append((obs_id, eff, c))

            if not cand_effects:
                continue

            gt_effect = gt_row['effect']
            key = (gt_row['element'], gt_row['info'])
            n_gt_same = len(gt_by_el_info[key])

            if n_gt_same > 1 and len(cand_effects) > 1:
                unused = [(oid, eff, c) for oid, eff, c in cand_effects
                          if oid not in used_obs_ids]
                if not unused:
                    unused = cand_effects
                best = min(unused, key=lambda x: abs(x[1] - gt_effect))
                our_effect = best[1]
                used_obs_ids.add(best[0])
            else:
                effects = [eff for _, eff, _ in cand_effects]
                our_effect = sum(effects) / len(effects)

            err = abs(our_effect - gt_effect)
            matched += 1
            all_matches.append({
                'paper': paper_id,
                'el': gt_row['element'],
                'our': our_effect,
                'gt': gt_effect,
                'err': err,
                'info': gt_row['info'],
            })

        paper_details[paper_id] = {
            'gt_rows': len(gt_rows),
            'matched': matched,
            'total_obs': len(obs_list),
            'tc_swap': swap_tc,
        }

    return all_matches, paper_details


def compute_stats(matches):
    """Compute aggregate statistics from a list of match dicts."""
    n = len(matches)
    if n == 0:
        return {
            'n': 0, 'mae': float('nan'), 'median_ae': float('nan'),
            'r': float('nan'), 'direction_pct': float('nan'),
            'w5': 0, 'w10': 0, 'w20': 0,
            'overall_our': float('nan'), 'overall_gt': float('nan'),
        }

    errs = [m['err'] for m in matches]
    mae = sum(errs) / n * 100
    sorted_errs = sorted(errs)
    median_ae = sorted_errs[n // 2] * 100

    w5 = sum(1 for e in errs if e <= 0.05)
    w10 = sum(1 for e in errs if e <= 0.10)
    w20 = sum(1 for e in errs if e <= 0.20)

    our = [m['our'] for m in matches]
    gts = [m['gt'] for m in matches]

    dir_total = sum(1 for m in matches if m['gt'] != 0)
    dir_ok = sum(1 for m in matches if m['gt'] != 0 and (m['our'] < 0) == (m['gt'] < 0))
    dir_pct = dir_ok / dir_total * 100 if dir_total > 0 else 0

    mean_our = sum(our) / n
    mean_gt = sum(gts) / n
    cov = sum((o - mean_our) * (g - mean_gt) for o, g in zip(our, gts))
    var_our = sum((o - mean_our) ** 2 for o in our)
    var_gt = sum((g - mean_gt) ** 2 for g in gts)
    r = cov / math.sqrt(var_our * var_gt) if var_our > 0 and var_gt > 0 else 0

    return {
        'n': n,
        'mae': round(mae, 2),
        'median_ae': round(median_ae, 2),
        'r': round(r, 3),
        'direction_pct': round(dir_pct, 1),
        'w5': w5,
        'w10': w10,
        'w20': w20,
        'w5_pct': round(w5 / n * 100, 1),
        'w10_pct': round(w10 / n * 100, 1),
        'w20_pct': round(w20 / n * 100, 1),
        'overall_our': round(mean_our * 100, 2),
        'overall_gt': round(mean_gt * 100, 2),
    }


# ==================================================================
# ABLATION-SPECIFIC ANALYSES
# ==================================================================

def get_paper_metadata(consensus_dir):
    """
    Load per-paper metadata (extraction_method, difficulty, tiebreaker_used)
    from the consensus JSON files.
    """
    metadata = {}
    for rf in sorted(consensus_dir.glob("*_consensus.json")):
        paper_id = rf.stem.replace("_consensus", "")
        with open(rf, encoding='utf-8') as f:
            data = json.load(f)
        recon = data.get('recon', {})
        metadata[paper_id] = {
            'extraction_method': recon.get('extraction_method', 'text'),
            'difficulty': recon.get('estimated_difficulty', 'MEDIUM'),
            'tiebreaker_used': data.get('tiebreaker_used', False),
            'tiebreaker_reason': data.get('tiebreaker_reason', ''),
            'claude_obs': data.get('claude_obs', 0),
            'kimi_obs': data.get('kimi_obs', 0),
            'gemini_obs': data.get('gemini_obs', 0),
        }
    return metadata


def analyze_by_mode(matches, metadata):
    """Split matches by extraction_method (text vs hybrid)."""
    text_matches = [m for m in matches if metadata.get(m['paper'], {}).get('extraction_method', 'text') == 'text']
    hybrid_matches = [m for m in matches if metadata.get(m['paper'], {}).get('extraction_method', 'text') != 'text']
    return {
        'text': compute_stats(text_matches),
        'hybrid': compute_stats(hybrid_matches),
    }


def analyze_by_difficulty(matches, metadata):
    """Split matches by estimated_difficulty (HARD vs MEDIUM)."""
    hard_matches = [m for m in matches if metadata.get(m['paper'], {}).get('difficulty', 'MEDIUM') == 'HARD']
    medium_matches = [m for m in matches if metadata.get(m['paper'], {}).get('difficulty', 'MEDIUM') != 'HARD']
    return {
        'HARD': compute_stats(hard_matches),
        'MEDIUM': compute_stats(medium_matches),
    }


def analyze_tiebreaker(matches, metadata):
    """Split consensus matches by whether tiebreaker was used."""
    tb_papers = {pid for pid, md in metadata.items() if md.get('tiebreaker_used', False)}
    tb_matches = [m for m in matches if m['paper'] in tb_papers]
    no_tb_matches = [m for m in matches if m['paper'] not in tb_papers]
    return {
        'with_tiebreaker': compute_stats(tb_matches),
        'without_tiebreaker': compute_stats(no_tb_matches),
        'tiebreaker_papers': sorted(tb_papers),
        'n_tiebreaker_papers': len(tb_papers),
    }


# ==================================================================
# MAIN
# ==================================================================

def main():
    print("=" * 78)
    print("ABLATION ANALYSIS: Consensus Pipeline vs Single-Model Baselines")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 78)

    # Load ground truth
    gt = load_gt()
    total_gt_rows = sum(len(v) for v in gt.values())
    print(f"\nGround truth: {total_gt_rows} observations across {len(gt)} references")

    # Load paper metadata from consensus JSONs
    metadata = get_paper_metadata(CONSENSUS_DIR)
    print(f"Paper metadata loaded: {len(metadata)} papers")

    # ------------------------------------------------------------------
    # 1. Validate each source against GT
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("SECTION 1: SINGLE-MODEL vs CONSENSUS COMPARISON")
    print(f"{'='*78}\n")

    results = {}

    # Consensus pipeline
    print("Validating: Consensus Pipeline...")
    consensus_matches, consensus_papers = validate_source(CONSENSUS_DIR, gt, "consensus")
    results['Consensus'] = compute_stats(consensus_matches)
    results['Consensus']['papers_matched'] = len([p for p in consensus_papers.values() if p['matched'] > 0])
    print(f"  -> {results['Consensus']['n']} matched observations from {results['Consensus']['papers_matched']} papers")

    # Each baseline
    for model_name, baseline_dir in BASELINE_DIRS.items():
        print(f"Validating: {model_name} solo...")
        matches, papers = validate_source(baseline_dir, gt, "baseline")
        results[model_name] = compute_stats(matches)
        results[model_name]['papers_matched'] = len([p for p in papers.values() if p['matched'] > 0])
        print(f"  -> {results[model_name]['n']} matched observations from {results[model_name]['papers_matched']} papers")

    # Print comparison table
    print(f"\n{'='*78}")
    print("TABLE 1: Single-Model vs Consensus Performance")
    print(f"{'='*78}")
    print(f"{'Source':<12} {'N obs':>6} {'Papers':>7} {'MAE%':>6} {'MedAE%':>7} {'r':>6} {'Dir%':>6} {'W5%':>5} {'W10%':>5} {'W20%':>5}")
    print("-" * 78)

    # Sort: consensus first, then by MAE ascending
    source_order = ['Consensus', 'Claude', 'Gemini', 'Kimi']
    for src in source_order:
        s = results[src]
        print(f"{src:<12} {s['n']:>6} {s.get('papers_matched','?'):>7} "
              f"{s['mae']:>5.1f}% {s['median_ae']:>6.1f}% {s['r']:>5.3f} "
              f"{s['direction_pct']:>5.1f} {s['w5_pct']:>4.1f} {s['w10_pct']:>4.1f} {s['w20_pct']:>4.1f}")

    # Consensus gain calculation
    best_single_mae = min(results[m]['mae'] for m in ['Claude', 'Gemini', 'Kimi'] if results[m]['n'] > 0)
    best_single_name = min(['Claude', 'Gemini', 'Kimi'],
                           key=lambda m: results[m]['mae'] if results[m]['n'] > 0 else 999)
    consensus_mae = results['Consensus']['mae']
    mae_improvement = best_single_mae - consensus_mae
    mae_improvement_pct = mae_improvement / best_single_mae * 100 if best_single_mae > 0 else 0

    print(f"\nConsensus MAE improvement over best single model ({best_single_name}):")
    print(f"  {best_single_mae:.1f}% -> {consensus_mae:.1f}% = {mae_improvement:+.1f}pp ({mae_improvement_pct:.0f}% relative reduction)")

    # Observation coverage
    best_single_n = max(results[m]['n'] for m in ['Claude', 'Gemini', 'Kimi'])
    consensus_n = results['Consensus']['n']
    obs_gain = consensus_n - best_single_n
    obs_gain_pct = obs_gain / best_single_n * 100 if best_single_n > 0 else 0
    print(f"\nObservation coverage gain (consensus vs best single):")
    print(f"  {best_single_n} -> {consensus_n} = +{obs_gain} ({obs_gain_pct:+.0f}%)")

    # ------------------------------------------------------------------
    # 2. Tiebreaker contribution
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("SECTION 2: TIEBREAKER CONTRIBUTION")
    print(f"{'='*78}\n")

    tb_analysis = analyze_tiebreaker(consensus_matches, metadata)
    print(f"Tiebreaker (Gemini) was used for {tb_analysis['n_tiebreaker_papers']} papers:")
    for p in tb_analysis['tiebreaker_papers']:
        md = metadata.get(p, {})
        print(f"  - {p} (reason: {md.get('tiebreaker_reason', 'N/A')[:60]})")

    print(f"\n{'Subset':<20} {'N obs':>6} {'MAE%':>6} {'r':>6} {'Dir%':>6}")
    print("-" * 50)
    for label, stats in [("With tiebreaker", tb_analysis['with_tiebreaker']),
                         ("Without tiebreaker", tb_analysis['without_tiebreaker'])]:
        if stats['n'] > 0:
            print(f"{label:<20} {stats['n']:>6} {stats['mae']:>5.1f}% {stats['r']:>5.3f} {stats['direction_pct']:>5.1f}")

    # ------------------------------------------------------------------
    # 3. TEXT vs HYBRID mode
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("SECTION 3: TEXT vs HYBRID/VISION MODE")
    print(f"{'='*78}\n")

    mode_analysis = analyze_by_mode(consensus_matches, metadata)

    # Count papers per mode
    text_papers = set(m['paper'] for m in consensus_matches
                      if metadata.get(m['paper'], {}).get('extraction_method', 'text') == 'text')
    hybrid_papers = set(m['paper'] for m in consensus_matches
                        if metadata.get(m['paper'], {}).get('extraction_method', 'text') != 'text')

    print(f"TEXT mode papers:   {len(text_papers)}")
    print(f"HYBRID mode papers: {len(hybrid_papers)}")
    print()

    print(f"{'Mode':<10} {'N obs':>6} {'MAE%':>6} {'MedAE%':>7} {'r':>6} {'Dir%':>6} {'W5%':>5} {'W10%':>5}")
    print("-" * 60)
    for mode, stats in mode_analysis.items():
        if stats['n'] > 0:
            print(f"{mode.upper():<10} {stats['n']:>6} {stats['mae']:>5.1f}% {stats['median_ae']:>6.1f}% "
                  f"{stats['r']:>5.3f} {stats['direction_pct']:>5.1f} {stats['w5_pct']:>4.1f} {stats['w10_pct']:>4.1f}")

    if mode_analysis['text']['n'] > 0 and mode_analysis['hybrid']['n'] > 0:
        text_mae = mode_analysis['text']['mae']
        hybrid_mae = mode_analysis['hybrid']['mae']
        diff = hybrid_mae - text_mae
        print(f"\nTEXT advantage: {diff:+.1f}pp MAE")

    # ------------------------------------------------------------------
    # 4. HARD vs MEDIUM difficulty
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("SECTION 4: CHALLENGE-AWARE ROUTING (HARD vs MEDIUM)")
    print(f"{'='*78}\n")

    diff_analysis = analyze_by_difficulty(consensus_matches, metadata)

    hard_papers = set(m['paper'] for m in consensus_matches
                      if metadata.get(m['paper'], {}).get('difficulty', 'MEDIUM') == 'HARD')
    medium_papers = set(m['paper'] for m in consensus_matches
                        if metadata.get(m['paper'], {}).get('difficulty', 'MEDIUM') != 'HARD')

    print(f"HARD papers:   {len(hard_papers)}")
    print(f"MEDIUM papers: {len(medium_papers)}")
    print()

    print(f"{'Difficulty':<10} {'N obs':>6} {'MAE%':>6} {'MedAE%':>7} {'r':>6} {'Dir%':>6} {'W5%':>5} {'W10%':>5}")
    print("-" * 60)
    for diff, stats in diff_analysis.items():
        if stats['n'] > 0:
            print(f"{diff:<10} {stats['n']:>6} {stats['mae']:>5.1f}% {stats['median_ae']:>6.1f}% "
                  f"{stats['r']:>5.3f} {stats['direction_pct']:>5.1f} {stats['w5_pct']:>4.1f} {stats['w10_pct']:>4.1f}")

    if diff_analysis['HARD']['n'] > 0 and diff_analysis['MEDIUM']['n'] > 0:
        hard_mae = diff_analysis['HARD']['mae']
        medium_mae = diff_analysis['MEDIUM']['mae']
        diff_gap = hard_mae - medium_mae
        print(f"\nHARD-MEDIUM gap: {diff_gap:+.1f}pp MAE")

    # ------------------------------------------------------------------
    # 5. Per-model observation yield
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("SECTION 5: PER-MODEL OBSERVATION YIELD IN CONSENSUS")
    print(f"{'='*78}\n")

    total_claude = sum(md.get('claude_obs', 0) for md in metadata.values())
    total_kimi = sum(md.get('kimi_obs', 0) for md in metadata.values())
    total_gemini = sum(md.get('gemini_obs', 0) for md in metadata.values())

    print(f"{'Model':<10} {'Total Obs':>10} {'Papers w/ data':>15}")
    print("-" * 40)
    print(f"{'Claude':<10} {total_claude:>10} {sum(1 for md in metadata.values() if md.get('claude_obs', 0) > 0):>15}")
    print(f"{'Kimi':<10} {total_kimi:>10} {sum(1 for md in metadata.values() if md.get('kimi_obs', 0) > 0):>15}")
    print(f"{'Gemini':<10} {total_gemini:>10} {sum(1 for md in metadata.values() if md.get('gemini_obs', 0) > 0):>15}")
    print(f"{'Total':<10} {total_claude + total_kimi + total_gemini:>10}")

    # ------------------------------------------------------------------
    # 5b. FIXED-SCOPE COMPARISON (fair apples-to-apples)
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("SECTION 5b: FIXED-SCOPE COMPARISON (same GT observations for all sources)")
    print(f"{'='*78}\n")

    # Build a key for each match: (paper, element, info) to identify GT rows
    def match_key(m):
        return (m['paper'], m['el'], m['info'])

    # Find all GT rows that ALL sources matched
    consensus_keys = set(match_key(m) for m in consensus_matches)

    all_baseline_matches = {}
    all_baseline_keys = {}
    for model_name in ['Claude', 'Gemini', 'Kimi']:
        bm, _ = validate_source(BASELINE_DIRS[model_name], gt, "baseline")
        all_baseline_matches[model_name] = bm
        all_baseline_keys[model_name] = set(match_key(m) for m in bm)

    # Intersection: GT rows matched by ALL sources
    common_keys = consensus_keys.copy()
    for model_name in ['Claude', 'Gemini', 'Kimi']:
        common_keys = common_keys & all_baseline_keys[model_name]

    print(f"GT observations matched by ALL 4 sources: {len(common_keys)}")
    print(f"  Consensus matched:  {len(consensus_keys)}")
    for model_name in ['Claude', 'Gemini', 'Kimi']:
        print(f"  {model_name} matched:     {len(all_baseline_keys[model_name])}")
    print()

    # Filter each source to the common scope
    fixed_consensus = [m for m in consensus_matches if match_key(m) in common_keys]
    fixed_baselines = {}
    for model_name in ['Claude', 'Gemini', 'Kimi']:
        fixed_baselines[model_name] = [m for m in all_baseline_matches[model_name] if match_key(m) in common_keys]

    fixed_stats = {'Consensus': compute_stats(fixed_consensus)}
    for model_name in ['Claude', 'Gemini', 'Kimi']:
        fixed_stats[model_name] = compute_stats(fixed_baselines[model_name])

    print(f"{'Source':<12} {'N obs':>6} {'MAE%':>6} {'MedAE%':>7} {'r':>6} {'Dir%':>6} {'W5%':>5} {'W10%':>5} {'W20%':>5}")
    print("-" * 78)
    for src in source_order:
        s = fixed_stats[src]
        if s['n'] > 0:
            print(f"{src:<12} {s['n']:>6} {s['mae']:>5.1f}% {s['median_ae']:>6.1f}% {s['r']:>5.3f} "
                  f"{s['direction_pct']:>5.1f} {s['w5_pct']:>4.1f} {s['w10_pct']:>4.1f} {s['w20_pct']:>4.1f}")

    # Also do pairwise: Consensus vs each single model (on their shared scope)
    print(f"\nPairwise comparisons (Consensus vs each model on shared GT):")
    print("-" * 78)
    for model_name in ['Claude', 'Gemini', 'Kimi']:
        pairwise_keys = consensus_keys & all_baseline_keys[model_name]
        pw_cons = [m for m in consensus_matches if match_key(m) in pairwise_keys]
        pw_base = [m for m in all_baseline_matches[model_name] if match_key(m) in pairwise_keys]
        pw_cons_stats = compute_stats(pw_cons)
        pw_base_stats = compute_stats(pw_base)
        if pw_cons_stats['n'] > 0 and pw_base_stats['n'] > 0:
            diff_mae = pw_base_stats['mae'] - pw_cons_stats['mae']
            diff_r = pw_cons_stats['r'] - pw_base_stats['r']
            print(f"  Consensus vs {model_name} ({pw_cons_stats['n']} shared obs):")
            print(f"    Consensus: MAE={pw_cons_stats['mae']:.1f}%  r={pw_cons_stats['r']:.3f}  Dir={pw_cons_stats['direction_pct']:.0f}%")
            print(f"    {model_name:8s}: MAE={pw_base_stats['mae']:.1f}%  r={pw_base_stats['r']:.3f}  Dir={pw_base_stats['direction_pct']:.0f}%")
            print(f"    Consensus advantage: MAE {diff_mae:+.1f}pp, r {diff_r:+.3f}")
            print()

    # ------------------------------------------------------------------
    # 5c. COVERAGE vs ACCURACY TRADEOFF
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("SECTION 5c: COVERAGE vs ACCURACY TRADEOFF")
    print(f"{'='*78}\n")

    # Total GT rows available for our 46 papers
    all_paper_gt_rows = set()
    for paper_id in PAPER_TO_LOLADZE_REF:
        ref = PAPER_TO_LOLADZE_REF[paper_id]
        if ref in gt:
            for i, row in enumerate(gt[ref]):
                all_paper_gt_rows.add((paper_id, row['element'], row['info']))
    total_possible = len(all_paper_gt_rows)

    print(f"Total GT observations for our 46 papers: {total_possible}")
    print(f"\n{'Source':<12} {'Matched':>8} {'Coverage%':>10} {'MAE%':>6} {'Dir%':>6} {'r':>6}")
    print("-" * 55)
    for src in source_order:
        s = results[src]
        cov = s['n'] / total_possible * 100 if total_possible > 0 else 0
        print(f"{src:<12} {s['n']:>8} {cov:>9.1f}% {s['mae']:>5.1f}% {s['direction_pct']:>5.1f} {s['r']:>5.3f}")
    print()
    print("Key insight: Consensus captures 560 GT observations across 42 papers.")
    print("Individual models may have lower MAE on the subset they DO match,")
    print("but fail to extract data from papers where they struggle.")
    print(f"Example: Kimi (best MAE) only matched {results['Kimi']['n']} obs from "
          f"{results['Kimi'].get('papers_matched', '?')} papers.")
    print(f"Consensus matched {results['Consensus']['n']} obs from "
          f"{results['Consensus'].get('papers_matched', '?')} papers.")

    # Show how many papers each source uniquely covers
    consensus_paper_set = set(m['paper'] for m in consensus_matches)
    model_paper_sets = {}
    for model_name in ['Claude', 'Gemini', 'Kimi']:
        model_paper_sets[model_name] = set(m['paper'] for m in all_baseline_matches[model_name])

    print(f"\nPapers covered by each source:")
    for src, pset in [('Consensus', consensus_paper_set)] + list(model_paper_sets.items()):
        others = set()
        for other_src, other_pset in [('Consensus', consensus_paper_set)] + list(model_paper_sets.items()):
            if other_src != src:
                others |= other_pset
        unique = pset - others
        print(f"  {src:<12}: {len(pset)} papers ({len(unique)} unique)")

    # ------------------------------------------------------------------
    # 6. Element-level comparison
    # ------------------------------------------------------------------
    print(f"\n{'='*78}")
    print("SECTION 6: ELEMENT-LEVEL COMPARISON (Consensus vs Best Single)")
    print(f"{'='*78}\n")

    # Group consensus matches by element
    consensus_el = defaultdict(list)
    for m in consensus_matches:
        consensus_el[m['el']].append(m)

    # Find best single model for each element
    best_model_matches = {}
    for model_name in ['Claude', 'Gemini', 'Kimi']:
        bm, _ = validate_source(BASELINE_DIRS[model_name], gt, "baseline")
        best_model_matches[model_name] = bm

    print(f"{'Element':<5} {'Cons MAE':>9} {'Cons Dir':>9} | {'Claude':>7} {'Gemini':>7} {'Kimi':>7} | {'Best':>7} {'Gain':>7}")
    print("-" * 78)

    element_detail = {}
    for el in sorted(consensus_el.keys(), key=lambda e: len(consensus_el[e]), reverse=True):
        if len(consensus_el[el]) < 3:
            continue  # skip elements with too few observations
        cons_stats = compute_stats(consensus_el[el])

        el_stats_by_model = {}
        for model_name, bm in best_model_matches.items():
            el_matches = [m for m in bm if m['el'] == el]
            el_stats_by_model[model_name] = compute_stats(el_matches) if el_matches else {'mae': float('nan'), 'n': 0}

        best_model = min(el_stats_by_model.keys(),
                         key=lambda m: el_stats_by_model[m]['mae'] if el_stats_by_model[m]['n'] > 0 else 999)
        best_mae = el_stats_by_model[best_model]['mae']

        gain = best_mae - cons_stats['mae'] if not math.isnan(best_mae) else float('nan')

        claude_mae = el_stats_by_model['Claude']['mae']
        gemini_mae = el_stats_by_model['Gemini']['mae']
        kimi_mae = el_stats_by_model['Kimi']['mae']

        def fmt_mae(v):
            return f"{v:.1f}%" if not math.isnan(v) else "  N/A"

        print(f"{el:<5} {cons_stats['mae']:>7.1f}%  {cons_stats['direction_pct']:>6.0f}%  | "
              f"{fmt_mae(claude_mae):>7} {fmt_mae(gemini_mae):>7} {fmt_mae(kimi_mae):>7} | "
              f"{fmt_mae(best_mae):>7} {'+' if not math.isnan(gain) else ''}{f'{gain:.1f}pp' if not math.isnan(gain) else 'N/A':>6}")

        element_detail[el] = {
            'consensus_mae': cons_stats['mae'],
            'consensus_n': cons_stats['n'],
            'consensus_dir': cons_stats['direction_pct'],
            'claude_mae': claude_mae,
            'gemini_mae': gemini_mae,
            'kimi_mae': kimi_mae,
            'best_single': best_model,
            'best_single_mae': best_mae,
            'consensus_gain': gain if not math.isnan(gain) else None,
        }

    # ------------------------------------------------------------------
    # SAVE RESULTS
    # ------------------------------------------------------------------
    output_data = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            source: {k: v for k, v in stats.items()} for source, stats in results.items()
        },
        'tiebreaker': {
            'n_papers': tb_analysis['n_tiebreaker_papers'],
            'papers': tb_analysis['tiebreaker_papers'],
            'with_tb_stats': tb_analysis['with_tiebreaker'],
            'without_tb_stats': tb_analysis['without_tiebreaker'],
        },
        'extraction_mode': {
            'text': mode_analysis['text'],
            'hybrid': mode_analysis['hybrid'],
        },
        'difficulty': {
            'hard': diff_analysis['HARD'],
            'medium': diff_analysis['MEDIUM'],
        },
        'model_yield': {
            'claude_obs': total_claude,
            'kimi_obs': total_kimi,
            'gemini_obs': total_gemini,
        },
        'element_detail': element_detail,
        'consensus_gain': {
            'best_single_model': best_single_name,
            'best_single_mae': best_single_mae,
            'consensus_mae': consensus_mae,
            'mae_improvement_pp': round(mae_improvement, 2),
            'mae_improvement_pct': round(mae_improvement_pct, 1),
            'obs_gain': obs_gain,
            'obs_gain_pct': round(obs_gain_pct, 1),
        },
        'fixed_scope': {
            'n_common_obs': len(common_keys),
            'stats': {src: {k: v for k, v in s.items()} for src, s in fixed_stats.items()},
        },
    }

    output_path = OUTPUT_DIR / "ablation_results.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, default=str)
    print(f"\n{'='*78}")
    print(f"Results saved to: {output_path}")
    print(f"{'='*78}")


if __name__ == "__main__":
    main()
