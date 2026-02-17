"""
Replicate the Loladze 2014 meta-analysis from each extraction source.

The GT meta-analysis answers: "Does elevated CO2 reduce plant mineral concentrations?"
- Scope: 27 mineral elements, 46 papers
- Metric: (E-A)/A as percent change

For each extraction source, we:
1. Filter to ONLY mineral element observations (matching the 27 GT elements)
2. Compute delta% = (treatment - control) / control * 100
3. Run random-effects meta-analysis
4. Compare to GT meta-analysis results

This is the FAIR comparison — same scope, same question, different data sources.
"""
import sys, os, json, math
from pathlib import Path
from collections import defaultdict

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import numpy as np
from scipy import stats

BASE_DIR = Path(r"C:\Users\moshe\Dropbox\Testing metaanalyis program")
EXTRACTOR_DIR = BASE_DIR / "meta_analysis_extractor"

# The 27 elements in Loladze GT
GT_ELEMENTS = {
    "Al", "B", "Ba", "Br", "C", "Ca", "Cd", "Cl", "Co", "Cr", "Cu",
    "Fe", "K", "Mg", "Mn", "Mo", "N", "Na", "Ni", "P", "Pb", "S",
    "Se", "Si", "Sr", "V", "Zn"
}

ELEMENT_ALIASES = {
    "calcium": "Ca", "ca": "Ca", "nitrogen": "N", "n": "N",
    "phosphorus": "P", "p": "P", "potassium": "K", "k": "K",
    "magnesium": "Mg", "mg": "Mg", "manganese": "Mn", "mn": "Mn",
    "iron": "Fe", "fe": "Fe", "zinc": "Zn", "zn": "Zn",
    "copper": "Cu", "cu": "Cu", "sulfur": "S", "sulphur": "S", "s": "S",
    "sodium": "Na", "na": "Na", "boron": "B", "b": "B",
    "molybdenum": "Mo", "mo": "Mo", "carbon": "C", "c": "C",
    "cobalt": "Co", "co": "Co", "aluminum": "Al", "aluminium": "Al", "al": "Al",
    "cadmium": "Cd", "cd": "Cd", "lead": "Pb", "pb": "Pb",
    "nickel": "Ni", "ni": "Ni", "vanadium": "V", "v": "V",
    "selenium": "Se", "se": "Se", "silicon": "Si", "si": "Si",
    "strontium": "Sr", "sr": "Sr", "barium": "Ba", "ba": "Ba",
    "chromium": "Cr", "cr": "Cr", "chlorine": "Cl", "cl": "Cl",
    "bromine": "Br", "br": "Br",
    # Common LLM outputs
    "total nitrogen": "N", "total n": "N", "total p": "P",
    "total phosphorus": "P", "total carbon": "C",
}


def normalize_element(name):
    if not name:
        return None
    clean = name.strip().lower()
    result = ELEMENT_ALIASES.get(clean)
    if result:
        return result
    # Try capitalizing (Fe, Zn, Ca, etc.)
    cap = name.strip().capitalize() if len(name.strip()) > 1 else name.strip().upper()
    if cap in GT_ELEMENTS:
        return cap
    up = name.strip().upper()
    if up in GT_ELEMENTS:
        return up
    return None  # Not a recognized mineral element


def random_effects_meta(effects):
    """Unweighted random-effects meta-analysis."""
    effects = np.array([e for e in effects if e is not None and np.isfinite(e)])
    k = len(effects)
    if k == 0:
        return {"pooled": None, "ci_lo": None, "ci_hi": None, "k": 0, "se": None}
    if k == 1:
        return {"pooled": float(effects[0]), "ci_lo": float(effects[0]),
                "ci_hi": float(effects[0]), "k": 1, "se": 0}
    mean_eff = float(np.mean(effects))
    se = float(np.std(effects, ddof=1) / np.sqrt(k))
    return {
        "pooled": mean_eff,
        "ci_lo": mean_eff - 1.96 * se,
        "ci_hi": mean_eff + 1.96 * se,
        "k": k,
        "se": se,
        "median": float(np.median(effects)),
    }


# ============================================================
# Load GT
# ============================================================

def load_gt():
    """Load Loladze GT, filtered to 46 papers, return per-obs deltas."""
    import openpyxl

    VALID_REFS = {
        "Fernando et al 2012a", "Fernando et al 2012b", "Ziska et al 1997",
        "Baslam et al 2012", "Finzi et al 2001", "Niinemets et al 1999",
        "Azam et al 2013", "Woodin et al 1992", "Cambell & Sage 2002",
        "Barnes & Pffirrman 1992", "Hogy et al 2009", "Huluka et al 1994",
        "Wu et al 2004", "Keutgen & Chen 2001", "Lieffering et al 2004",
        "Pleijel & Danielsson 2009", "Fangmeier et al  2002",
        "Al-Rawahy et al 2013", "Baxter et al 1994", "Overdieck 1993",
        "Wilsey et al 1994", "Blank et al 2011", "Guo et al 2013",
        "Seneweera & Conroy 1997", "Peet et al 1986", "Mishra et al 2011",
        "Pal et al 2004", "Kanowski 2001", "Johnson et al 2003",
        "Oksanen et al 2005", "Schenk et al 1997", "Haase et al 2008",
        "Newbery 1995", "Heagle et al 1993", "Pfirrmann et al 1996",
        "Mjwara et al 1996", "Luomala et al 2005", "Natali et al 2009",
        "Housman et al 2012", "Porter & Grodzinski 1984",
        "Rodenkirchen et al 2009", "Khan et al 2012", "Singh et al 2013",
        "Polley et al 2011", "Niu et al 2013", "O'Neill et al 1987",
    }

    gt_path = BASE_DIR / "Loladze" / "CO2+Dataset.xlsx"
    wb = openpyxl.load_workbook(gt_path, data_only=True)
    ws = wb.active

    observations = []
    for row_idx in range(3, ws.max_row + 1):
        ref = ws.cell(row=row_idx, column=17).value
        element = ws.cell(row=row_idx, column=19).value
        delta = ws.cell(row=row_idx, column=20).value
        tissue = ws.cell(row=row_idx, column=5).value

        if not ref or str(ref) not in VALID_REFS:
            continue
        if element is None or delta is None:
            continue

        elem = normalize_element(str(element))
        if not elem:
            continue

        try:
            delta_pct = float(delta) * 100
        except (ValueError, TypeError):
            continue

        observations.append({
            "reference": str(ref),
            "element": elem,
            "tissue": str(tissue).lower().strip() if tissue else "unknown",
            "delta_pct": delta_pct,
        })

    wb.close()
    return observations


# ============================================================
# Load extraction sources (filtered to mineral elements only)
# ============================================================

def load_extraction(source_dir, source_name):
    """
    Load observations from an extraction directory.
    FILTER: only keep observations where element is a recognized mineral element.
    """
    source_path = Path(source_dir)
    if not source_path.exists():
        return []

    observations = []
    papers_loaded = set()
    papers_skipped = 0
    non_mineral_skipped = 0

    json_files = sorted(list(source_path.glob("*_consensus.json")) +
                        list(source_path.glob("*_baseline.json")))

    for json_file in json_files:
        paper_id = json_file.stem.replace("_consensus", "").replace("_baseline", "")

        try:
            with open(json_file, encoding='utf-8') as f:
                data = json.load(f)
        except Exception:
            continue

        if "error" in data and "observations" not in data and "consensus_observations" not in data:
            papers_skipped += 1
            continue

        obs_list = data.get("consensus_observations", data.get("observations", []))
        if not isinstance(obs_list, list):
            continue

        paper_obs = []
        for obs in obs_list:
            t_mean = obs.get("treatment_mean")
            c_mean = obs.get("control_mean")
            if t_mean is None or c_mean is None:
                continue
            try:
                t_mean = float(t_mean)
                c_mean = float(c_mean)
            except (ValueError, TypeError):
                continue
            if c_mean == 0 or c_mean < 0 or t_mean < 0:
                continue

            # Normalize element — returns None if not a mineral element
            elem = normalize_element(obs.get("element", ""))
            if elem is None:
                non_mineral_skipped += 1
                continue

            delta_pct = (t_mean - c_mean) / c_mean * 100
            tissue = str(obs.get("tissue", "unknown")).lower().strip()

            paper_obs.append({
                "paper_id": paper_id,
                "element": elem,
                "tissue": tissue,
                "delta_pct": delta_pct,
                "treatment_mean": t_mean,
                "control_mean": c_mean,
            })

        if paper_obs:
            papers_loaded.add(paper_id)
            observations.extend(paper_obs)

    print(f"  {source_name}: {len(observations)} mineral obs from {len(papers_loaded)} papers "
          f"(skipped {non_mineral_skipped} non-mineral, {papers_skipped} failed papers)")
    return observations


# ============================================================
# Main comparison
# ============================================================

def fmt_ma(ma):
    if ma["pooled"] is None:
        return "N/A"
    return f"{ma['pooled']:+.2f}% [{ma['ci_lo']:+.2f}, {ma['ci_hi']:+.2f}] (k={ma['k']})"


def collapse_to_study_level(obs_list, group_key_fn):
    """
    Collapse observations to study-level effects by averaging within groups.
    group_key_fn(obs) should return a tuple key like (paper, element).
    Returns list of dicts with averaged delta_pct per group.
    """
    groups = defaultdict(list)
    for o in obs_list:
        key = group_key_fn(o)
        groups[key].append(o["delta_pct"])

    collapsed = []
    for key, deltas in groups.items():
        collapsed.append({
            "key": key,
            "delta_pct": float(np.mean(deltas)),
            "n_obs": len(deltas),
            "element": key[1] if len(key) > 1 else "all",
        })
    return collapsed


def build_gt_scope(gt_obs):
    """
    Build per-paper scope from GT: which elements and tissues were selected.
    Returns dict: paper_id -> {"elements": set, "tissues": set, "n_obs": int}
    """
    scope = defaultdict(lambda: {"elements": set(), "tissues": set(), "n_obs": 0})
    for o in gt_obs:
        ref = o["reference"]
        scope[ref]["elements"].add(o["element"])
        scope[ref]["tissues"].add(o["tissue"])
        scope[ref]["n_obs"] += 1
    return dict(scope)


# Map paper_id to GT reference
PAPER_TO_GT_REF = {
    "001_Ma_2007":          "Fernando et al 2012b",
    "002_Ziska_1997":       "Ziska et al 1997",
    "003_Baslam_2012":      "Baslam et al 2012",
    "004_Finzi_2001":       "Finzi et al 2001",
    "005_Niinemets_1999":   "Niinemets et al 1999",
    "006_Azam_2013":        "Azam et al 2013",
    "007_Woodin_1992":      "Woodin et al 1992",
    "008_Campbell_2002":    "Cambell & Sage 2002",
    "009_Barnes_1992":      "Barnes & Pffirrman 1992",
    "010_Li_2010":          "Hogy et al 2009",
    "011_Huluka_1994":      "Huluka et al 1994",
    "012_Wu_2004":          "Wu et al 2004",
    "013_Keutgen_2001":     "Keutgen & Chen 2001",
    "014_Lieffering_2004":  "Lieffering et al 2004",
    "015_Pleijel_2009":     "Pleijel & Danielsson 2009",
    "016_Fernando_2012a":   "Fernando et al 2012a",
    "017_Fangmeier_2002":   "Fangmeier et al  2002",
    "018_Al-Rawahy_2013":   "Al-Rawahy et al 2013",
    "019_Baxter_1994":      "Baxter et al 1994",
    "020_Overdieck_1993":   "Overdieck 1993",
    "021_Wilsey_1994":      "Wilsey et al 1994",
    "022_Blank_2011":       "Blank et al 2011",
    "025_Guo_2011":         "Guo et al 2013",
    "026_Seneweera_1997":   "Seneweera & Conroy 1997",
    "027_Peet_1986":        "Peet et al 1986",
    "028_Mishra_2011":      "Mishra et al 2011",
    "031_Pal_2003":         "Pal et al 2004",
    "032_Kanowski_2001":    "Kanowski 2001",
    "034_Johnson_1997":     "Johnson et al 2003",
    "035_Oksanen_2005":     "Oksanen et al 2005",
    "036_Schenk_1997":      "Schenk et al 1997",
    "037_de_2000":          "Haase et al 2008",
    "038_Newbery_1995":     "Newbery 1995",
    "039_Heagle_1993":      "Heagle et al 1993",
    "040_Pfirrmann_1996":   "Pfirrmann et al 1996",
    "041_Mjwara_1996":      "Mjwara et al 1996",
    "042_Luomala_2005":     "Luomala et al 2005",
    "043_Natali_2009":      "Natali et al 2009",
    "044_Housman_2012":     "Housman et al 2012",
    "046_Porter_1984":      "Porter & Grodzinski 1984",
    "047_Rodenkirchen_2009":"Rodenkirchen et al 2009",
    "048_Khan_2013":        "Khan et al 2012",
    "049_Singh_2013":       "Singh et al 2013",
    "050_Polley_2011":      "Polley et al 2011",
    "051_Niu_2013":         "Niu et al 2013",
    "058_ONeill_1987":      "O'Neill et al 1987",
}

# Tissue normalization for matching GT tissues to extracted tissues
TISSUE_NORM = {
    "leaf": "foliar", "leaves": "foliar", "foliar": "foliar", "foliage": "foliar",
    "needle": "needles", "needles": "needles",
    "grain": "grain", "seed": "seed", "seeds": "seed",
    "shoot": "shoots", "shoots": "shoots",
    "root": "roots", "roots": "roots",
    "stem": "stems", "stems": "stems",
    "fruit": "fruit", "fruits": "fruit",
    "tuber": "tuber", "tubers": "tuber",
    "above ground": "above ground", "aboveground": "above ground",
    "whole plant": "above ground", "whole": "above ground",
    "stover": "stover", "tillers": "tillers", "blades": "blades",
    "frond": "frond", "fronds": "frond",
    "seedlings": "seedlings", "seedling": "seedlings",
    "inner leaves": "foliar", "old leaves": "foliar", "young leaves": "foliar",
    "flag leaf": "foliar", "mature leaves": "foliar",
}


def tissue_matches_gt(extracted_tissue, gt_tissues):
    """Check if an extracted tissue matches any GT tissue for this paper."""
    if not extracted_tissue or extracted_tissue == "unknown":
        # Unknown tissue — allow it if there's only one GT tissue
        return len(gt_tissues) == 1

    ext_norm = TISSUE_NORM.get(extracted_tissue, extracted_tissue)

    for gt_t in gt_tissues:
        gt_norm = TISSUE_NORM.get(gt_t, gt_t)
        if ext_norm == gt_norm:
            return True
        # Fuzzy: "above ground" matches "shoots", "stems" etc for whole-plant studies
        if gt_norm == "above ground" and ext_norm in ("shoots", "stems", "foliar", "above ground"):
            return True
        if ext_norm == "above ground" and gt_norm in ("shoots", "stems", "foliar"):
            return True

    return False


def filter_by_gt_scope(obs_list, gt_scope):
    """
    Filter extracted observations to only include paper×element×tissue
    combinations that exist in the GT.
    """
    kept = []
    dropped_elem = 0
    dropped_tissue = 0

    for obs in obs_list:
        paper_id = obs.get("paper_id", "")
        gt_ref = PAPER_TO_GT_REF.get(paper_id, "")

        if gt_ref not in gt_scope:
            continue

        scope = gt_scope[gt_ref]

        # Element filter
        if obs["element"] not in scope["elements"]:
            dropped_elem += 1
            continue

        # Tissue filter
        if not tissue_matches_gt(obs.get("tissue", ""), scope["tissues"]):
            dropped_tissue += 1
            continue

        kept.append(obs)

    return kept, dropped_elem, dropped_tissue


def main():
    print("=" * 80)
    print("REPLICATING LOLADZE 2014 META-ANALYSIS FROM EACH SOURCE")
    print("  Scope: 27 mineral elements, 46 papers, elevated CO2 effects")
    print("  Filter: ONLY mineral elements, collapsed to 1 effect per paper x element")
    print("=" * 80)

    # Load GT
    gt_obs = load_gt()
    gt_elements = set(o["element"] for o in gt_obs)
    gt_papers = set(o["reference"] for o in gt_obs)
    print(f"\nGT: {len(gt_obs)} obs, {len(gt_papers)} papers, {len(gt_elements)} elements")

    # Collapse GT to paper × element level
    gt_collapsed = collapse_to_study_level(
        gt_obs, lambda o: (o["reference"], o["element"])
    )
    print(f"GT collapsed: {len(gt_collapsed)} paper×element effects")

    # Build GT scope
    gt_scope = build_gt_scope(gt_obs)
    print(f"GT scope: {len(gt_scope)} papers with specific element+tissue selections")

    # Load sources
    sources = {
        "Pipeline": EXTRACTOR_DIR / "output" / "loladze_full_46_v2",
        "Claude (simple)": EXTRACTOR_DIR / "output" / "baseline_claude_loladze",
        "Gemini (simple)": EXTRACTOR_DIR / "output" / "baseline_gemini_loladze",
        "Kimi (simple)": EXTRACTOR_DIR / "output" / "baseline_kimi_loladze",
    }

    all_data = {}       # raw mineral obs
    all_filtered = {}   # filtered to GT scope (same elements + tissues per paper)
    all_collapsed = {}  # collapsed: 1 per paper×element (from filtered)

    for name, path in sources.items():
        obs = load_extraction(path, name)
        if obs:
            all_data[name] = obs

            # Filter to GT scope
            filtered, drop_e, drop_t = filter_by_gt_scope(obs, gt_scope)
            all_filtered[name] = filtered
            print(f"  → GT-scoped: {len(filtered)} obs (dropped {drop_e} wrong-element, {drop_t} wrong-tissue)")

            collapsed = collapse_to_study_level(
                filtered, lambda o: (o["paper_id"], o["element"])
            )
            all_collapsed[name] = collapsed
            print(f"  → collapsed to {len(collapsed)} paper×element effects")

    # ========================================================
    # 1. OVERALL META-ANALYSIS (raw obs and collapsed)
    # ========================================================
    print("\n" + "=" * 80)
    print("1. OVERALL POOLED EFFECT: Does elevated CO2 reduce mineral concentrations?")
    print("=" * 80)

    # Raw (all mineral obs, no filtering)
    gt_ma_raw = random_effects_meta([o["delta_pct"] for o in gt_obs])
    print(f"\n  A) ALL mineral observations (no scope filter, no collapsing):")
    print(f"  {'Source':<20} {'Pooled effect':>32} {'Diff':>8}")
    print(f"  {'-'*62}")
    print(f"  {'GT (Loladze)':<20} {fmt_ma(gt_ma_raw):>32} {'---':>8}")
    for name, obs_list in all_data.items():
        ma = random_effects_meta([o["delta_pct"] for o in obs_list])
        diff = abs(ma["pooled"] - gt_ma_raw["pooled"]) if ma["pooled"] is not None else None
        diff_str = f"{diff:.1f}" if diff is not None else "N/A"
        print(f"  {name:<20} {fmt_ma(ma):>32} {diff_str:>7}pp")

    # GT-scoped (same elements + tissues per paper)
    print(f"\n  B) GT-SCOPED: only GT elements + tissues per paper, all obs:")
    print(f"  {'Source':<20} {'Pooled effect':>32} {'Diff':>8}")
    print(f"  {'-'*62}")
    print(f"  {'GT (Loladze)':<20} {fmt_ma(gt_ma_raw):>32} {'---':>8}")
    for name, filtered in all_filtered.items():
        ma = random_effects_meta([o["delta_pct"] for o in filtered])
        diff = abs(ma["pooled"] - gt_ma_raw["pooled"]) if ma["pooled"] is not None else None
        diff_str = f"{diff:.1f}" if diff is not None else "N/A"
        print(f"  {name:<20} {fmt_ma(ma):>32} {diff_str:>7}pp")

    # GT-scoped + collapsed (1 per paper × element)
    gt_ma = random_effects_meta([o["delta_pct"] for o in gt_collapsed])
    print(f"\n  C) GT-SCOPED + COLLAPSED to 1 effect per paper×element:")
    print(f"  {'Source':<20} {'Pooled effect':>32} {'Diff':>8}")
    print(f"  {'-'*62}")
    print(f"  {'GT (Loladze)':<20} {fmt_ma(gt_ma):>32} {'---':>8}")
    for name, collapsed in all_collapsed.items():
        ma = random_effects_meta([o["delta_pct"] for o in collapsed])
        diff = abs(ma["pooled"] - gt_ma["pooled"]) if ma["pooled"] is not None else None
        diff_str = f"{diff:.1f}" if diff is not None else "N/A"
        print(f"  {name:<20} {fmt_ma(ma):>32} {diff_str:>7}pp")

    # ========================================================
    # 2. PER-ELEMENT SUBGROUP ANALYSIS (collapsed paper×element)
    # ========================================================
    print("\n" + "=" * 80)
    print("2. PER-ELEMENT SUBGROUP EFFECTS (collapsed, elements with >=5 GT entries)")
    print("=" * 80)

    gt_by_elem = defaultdict(list)
    for o in gt_collapsed:
        gt_by_elem[o["element"]].append(o["delta_pct"])

    top_elements = sorted([e for e, v in gt_by_elem.items() if len(v) >= 5],
                          key=lambda e: len(gt_by_elem[e]), reverse=True)

    src_by_elem = {}
    for name, collapsed in all_collapsed.items():
        src_by_elem[name] = defaultdict(list)
        for o in collapsed:
            src_by_elem[name][o["element"]].append(o["delta_pct"])

    # Table header
    header = f"\n  {'Elem':<5} {'GT eff':>8} {'GT k':>5}"
    for name in all_collapsed.keys():
        short = name.split("(")[0].strip()[:8]
        header += f"  {short:>8} {'k':>4}"
    print(header)
    print(f"  {'-'*(len(header)-2)}")

    elem_errors = {name: [] for name in all_collapsed.keys()}

    for elem in top_elements:
        gt_eff = random_effects_meta(gt_by_elem[elem])
        line = f"  {elem:<5} {gt_eff['pooled']:>+8.1f} {gt_eff['k']:>5}"

        for name in all_collapsed.keys():
            src_eff = random_effects_meta(src_by_elem[name].get(elem, []))
            if src_eff["pooled"] is not None:
                diff = src_eff["pooled"] - gt_eff["pooled"]
                elem_errors[name].append(abs(diff))
                line += f"  {src_eff['pooled']:>+8.1f} {src_eff['k']:>4}"
            else:
                line += f"  {'N/A':>8} {0:>4}"

        print(line)

    # Element-level summary
    print(f"\n  Element-level MAE (avg |diff| from GT):")
    for name in all_collapsed.keys():
        errs = elem_errors[name]
        if errs:
            print(f"    {name:<20} {np.mean(errs):.2f}pp over {len(errs)} elements")

    # ========================================================
    # 3. DIRECTION AGREEMENT
    # ========================================================
    print("\n" + "=" * 80)
    print("3. DIRECTION AGREEMENT BY ELEMENT (collapsed)")
    print("=" * 80)

    n_agree = {name: 0 for name in all_collapsed.keys()}
    n_total = {name: 0 for name in all_collapsed.keys()}

    for elem in top_elements:
        gt_dir = "+" if np.mean(gt_by_elem[elem]) >= 0 else "-"
        for name in all_collapsed.keys():
            vals = src_by_elem[name].get(elem, [])
            if vals:
                src_dir = "+" if np.mean(vals) >= 0 else "-"
                n_total[name] += 1
                if src_dir == gt_dir:
                    n_agree[name] += 1

    print(f"\n  {'Source':<20} {'Agree':>7} {'Total':>7} {'Rate':>7}")
    print(f"  {'-'*43}")
    for name in all_collapsed.keys():
        rate = n_agree[name] / n_total[name] * 100 if n_total[name] > 0 else 0
        print(f"  {name:<20} {n_agree[name]:>7} {n_total[name]:>7} {rate:>6.0f}%")

    # ========================================================
    # 4. CORRELATION OF PER-ELEMENT EFFECTS (collapsed)
    # ========================================================
    print("\n" + "=" * 80)
    print("4. CORRELATION: Per-element effect sizes vs GT (collapsed)")
    print("=" * 80)

    for name in all_collapsed.keys():
        gt_vals = []
        src_vals = []
        for elem in top_elements:
            gt_eff = np.mean(gt_by_elem[elem])
            src_data = src_by_elem[name].get(elem, [])
            if src_data:
                gt_vals.append(gt_eff)
                src_vals.append(np.mean(src_data))

        if len(gt_vals) >= 3:
            r, p = stats.pearsonr(gt_vals, src_vals)
            mae = np.mean(np.abs(np.array(gt_vals) - np.array(src_vals)))
            print(f"  {name:<20} r={r:.3f} (p={p:.4f}), MAE={mae:.1f}pp, n={len(gt_vals)} elements")

    # ========================================================
    # 5. SCIENTIFIC CONCLUSIONS (collapsed)
    # ========================================================
    print("\n" + "=" * 80)
    print("5. SCIENTIFIC CONCLUSIONS — Would you reach the same findings?")
    print("   (Using collapsed paper×element effects)")
    print("=" * 80)

    print(f"\n  Loladze 2014 key finding: 'Elevated CO2 significantly reduces")
    print(f"  mineral concentrations in plants by ~5-10%'")
    print()

    all_conclusions = {}
    for name, collapsed in [("GT", gt_collapsed)] + list(all_collapsed.items()):
        deltas = [o["delta_pct"] for o in collapsed]
        ma = random_effects_meta(deltas)

        significant = ma["ci_hi"] < 0 if ma["pooled"] is not None and ma["pooled"] < 0 else False
        direction = "decline" if ma["pooled"] < 0 else "increase" if ma["pooled"] is not None else "?"

        conclusion = f"Significant {direction}" if significant else f"Non-significant {direction}"
        all_conclusions[name] = {
            "pooled": ma["pooled"],
            "significant": significant,
            "direction": direction,
            "conclusion": conclusion,
        }
        print(f"  {name:<20} {ma['pooled']:>+7.2f}%  CI [{ma['ci_lo']:>+7.2f}, {ma['ci_hi']:>+7.2f}]  → {conclusion}")

    gt_conclusion = all_conclusions["GT"]["conclusion"]
    matching = sum(1 for name, c in all_conclusions.items()
                   if name != "GT" and c["conclusion"] == gt_conclusion)
    total = len(all_conclusions) - 1
    print(f"\n  Consensus: {matching}/{total} sources reach same conclusion as GT")

    # Per-element rankings
    print(f"\n  Top declining elements (GT):")
    ranked_gt = sorted(top_elements, key=lambda e: np.mean(gt_by_elem[e]))
    for elem in ranked_gt[:5]:
        gt_eff = np.mean(gt_by_elem[elem])
        line = f"    {elem}: GT={gt_eff:+.1f}%"
        for name in all_collapsed.keys():
            vals = src_by_elem[name].get(elem, [])
            if vals:
                short = name.split("(")[0].strip()[:8]
                line += f"  {short}={np.mean(vals):+.1f}%"
        print(line)

    print(f"\n  Elements that increase under CO2 (GT):")
    for elem in reversed(ranked_gt):
        gt_eff = np.mean(gt_by_elem[elem])
        if gt_eff <= 0:
            break
        line = f"    {elem}: GT={gt_eff:+.1f}%"
        for name in all_collapsed.keys():
            vals = src_by_elem[name].get(elem, [])
            if vals:
                short = name.split("(")[0].strip()[:8]
                line += f"  {short}={np.mean(vals):+.1f}%"
        print(line)

    # ========================================================
    # 6. RANKING: Which source is closest to GT?
    # ========================================================
    print("\n" + "=" * 80)
    print("6. OVERALL RANKING — Which source best replicates the GT meta-analysis?")
    print("=" * 80)

    gt_overall = gt_ma["pooled"]
    scores = []
    for name in all_collapsed.keys():
        ma = random_effects_meta([o["delta_pct"] for o in all_collapsed[name]])
        overall_diff = abs(ma["pooled"] - gt_overall) if ma["pooled"] is not None else 999

        # Element-level correlation
        gt_v, src_v = [], []
        for elem in top_elements:
            g = gt_by_elem.get(elem, [])
            s = src_by_elem[name].get(elem, [])
            if g and s:
                gt_v.append(np.mean(g))
                src_v.append(np.mean(s))
        r_val = stats.pearsonr(gt_v, src_v)[0] if len(gt_v) >= 3 else 0
        elem_mae = np.mean(np.abs(np.array(gt_v) - np.array(src_v))) if gt_v else 999

        # Direction agreement
        dir_agree = n_agree.get(name, 0)
        dir_total = n_total.get(name, 1)
        dir_rate = dir_agree / dir_total if dir_total > 0 else 0

        scores.append({
            "name": name,
            "overall_diff": overall_diff,
            "element_r": r_val,
            "element_mae": elem_mae,
            "direction_rate": dir_rate,
            "k": ma["k"] if ma["pooled"] is not None else 0,
        })

    # Sort by overall diff (closest to GT first)
    scores.sort(key=lambda x: x["overall_diff"])

    print(f"\n  {'Rank':<5} {'Source':<20} {'Overall diff':>13} {'Elem r':>8} {'Elem MAE':>10} {'Dir agree':>10}")
    print(f"  {'-'*68}")
    for i, s in enumerate(scores, 1):
        print(f"  {i:<5} {s['name']:<20} {s['overall_diff']:>12.2f}pp {s['element_r']:>8.3f} {s['element_mae']:>9.2f}pp {s['direction_rate']:>9.0%}")


if __name__ == "__main__":
    main()
