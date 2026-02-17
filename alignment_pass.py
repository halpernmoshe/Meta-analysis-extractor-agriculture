"""
Post-Extraction Alignment Pass
================================
Uses an LLM to match extracted observations to ground truth metadata,
improving alignment without changing extraction.

The problem: validate_full_46.py has detailed filters for matching
extracted obs to GT rows (DOY, tissue, factorial level, etc.), but
the extraction often doesn't tag observations with the right moderator
keys. The filters exist but can't fire.

This script:
1. For each "misaligned" paper, takes ALL extracted observations
2. Takes the GT metadata (Additional Info, tissue, species, etc.)
3. Asks an LLM: "Which of these extracted observations best matches
   this GT description?"
4. Outputs a mapping file that validate_full_46.py can use

This is post-extraction alignment, NOT re-extraction. The numbers
stay the same; we just match them to the right GT rows.
"""

import sys, os, json, re, csv, math
from pathlib import Path
from collections import defaultdict
from datetime import datetime

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl

from dotenv import load_dotenv
load_dotenv()

BASE = Path(r"C:\Users\moshe\Dropbox\Testing metaanalyis program\meta_analysis_extractor")
GT_PATH = Path(r"C:\Users\moshe\Dropbox\Testing metaanalyis program\Loladze\CO2+Dataset.xlsx")
RESULTS_DIR = BASE / "output" / "loladze_full_46_v2"
OUTPUT_DIR = BASE / "output" / "alignment_pass"

# Papers with known alignment issues (from SPOT_CHECK_LOG.md)
ALIGNMENT_PAPERS = {
    "039_Heagle_1993": "tissue type mismatch — GT wants foliar, we extracted whole canopy",
    "011_Huluka_1994": "sampling date — GT wants DOY 247 (September), we extracted June (Table 2)",
    "040_Pfirrmann_1996": "factorial — GT separates +K/-K, we may have collapsed",
    "017_Fangmeier_2002": "factorial — GT separates CO2-only vs CO2+O3 conditions",
    "025_Guo_2011": "factorial over-extraction — GT wants Cd=0,Cu=0 subset only",
    "043_Natali_2009": "vision OCR + species matching at SERC site",
    "041_Mjwara_1996": "time course — GT wants final harvest (40 DAG), we extracted all timepoints",
    "050_Polley_2011": "gradient design — need to match specific CO2 levels to T/C",
    "014_Lieffering_2004": "figure reading — consensus dropped half the elements",
}

# Paper mapping
PAPER_TO_LOLADZE_REF = {
    "001_Ma_2007": "Fernando et al 2012a",
    "002_Ziska_1997": "Ziska et al 1997",
    "003_Baslam_2012": "Baslam et al 2012",
    "004_Finzi_2001": "Finzi et al 2001",
    "005_Niinemets_1999": "Niinemets et al 1999",
    "006_Azam_2013": "Azam et al 2013",
    "007_Woodin_1992": "Woodin et al 1992",
    "008_Campbell_2002": "Cambell & Sage 2002",
    "009_Barnes_1992": "Barnes & Pffirrman 1992",
    "010_Li_2010": "Hogy et al 2009",
    "011_Huluka_1994": "Huluka et al 1994",
    "012_Wu_2004": "Wu et al 2004",
    "013_Keutgen_2001": "Keutgen & Chen 2001",
    "014_Lieffering_2004": "Lieffering et al 2004",
    "015_Pleijel_2009": "Pleijel & Danielsson 2009",
    "016_Fernando_2012a": "Fernando et al 2012b",
    "017_Fangmeier_2002": "Fangmeier et al  2002",
    "018_Al-Rawahy_2013": "Al-Rawahy et al 2013",
    "019_Baxter_1994": "Baxter et al 1994",
    "020_Overdieck_1993": "Overdieck 1993",
    "021_Wilsey_1994": "Wilsey et al 1994",
    "022_Blank_2011": "Blank et al 2011",
    "025_Guo_2011": "Guo et al 2013",
    "026_Seneweera_1997": "Seneweera & Conroy 1997",
    "027_Peet_1986": "Peet et al 1986",
    "028_Mishra_2011": "Mishra et al 2011",
    "031_Pal_2003": "Pal et al 2004",
    "032_Kanowski_2001": "Kanowski 2001",
    "034_Johnson_1997": "Johnson et al 2003",
    "035_Oksanen_2005": "Oksanen et al 2005",
    "036_Schenk_1997": "Schenk et al 1997",
    "037_de_2000": "Haase et al 2008",
    "038_Newbery_1995": "Newbery 1995",
    "039_Heagle_1993": "Heagle et al 1993",
    "040_Pfirrmann_1996": "Pfirrmann et al 1996",
    "041_Mjwara_1996": "Mjwara et al 1996",
    "042_Luomala_2005": "Luomala et al 2005",
    "043_Natali_2009": "Natali et al 2009",
    "044_Housman_2012": "Housman et al 2012",
    "046_Porter_1984": "Porter & Grodzinski 1984",
    "047_Rodenkirchen_2009": "Rodenkirchen et al 2009",
    "048_Khan_2013": "Khan et al 2012",
    "049_Singh_2013": "Singh et al 2013",
    "050_Polley_2011": "Polley et al 2011",
    "051_Niu_2013": "Niu et al 2013",
    "058_ONeill_1987": "O'Neill et al 1987",
}


def normalize_element(el):
    if not el:
        return ""
    el = el.strip().upper()
    MAP = {"NITROGEN": "N", "PHOSPHORUS": "P", "POTASSIUM": "K",
           "CALCIUM": "CA", "MAGNESIUM": "MG", "IRON": "FE",
           "ZINC": "ZN", "MANGANESE": "MN", "COPPER": "CU",
           "SULFUR": "S", "SULPHUR": "S", "BORON": "B", "CARBON": "C",
           "SODIUM": "NA", "MOLYBDENUM": "MO", "SILICON": "SI",
           "SELENIUM": "SE", "COBALT": "CO", "CHROMIUM": "CR",
           "NICKEL": "NI", "LEAD": "PB", "VANADIUM": "V",
           "ALUMINUM": "AL", "ALUMINIUM": "AL", "BARIUM": "BA",
           "STRONTIUM": "SR", "CADMIUM": "CD", "PROTEIN": "PROTEIN"}
    return MAP.get(el, el)


def load_gt():
    """Load Loladze GT."""
    wb = openpyxl.load_workbook(GT_PATH, data_only=True)
    ws = wb["CO2 Dataset"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[2]]
    col = {}
    for i, h in enumerate(headers):
        if h == 'Reference': col['ref'] = i
        elif h == '(E-A)/A': col['effect'] = i
        elif h == 'Element': col['element'] = i
        elif h == 'Additional Info': col['info'] = i
        elif h == 'eCO2': col['eco2'] = i
        elif h == 'Tissue': col['tissue'] = i
        elif h == 'Species': col['species'] = i

    gt = defaultdict(list)
    for row in ws.iter_rows(min_row=3, values_only=True):
        ref = str(row[col['ref']]).strip() if row[col['ref']] else ""
        el = normalize_element(str(row[col['element']]).strip() if row[col['element']] else "")
        eff = row[col['effect']]
        info = str(row[col.get('info', 11)]).strip() if row[col.get('info', 11)] else ""
        eco2 = row[col.get('eco2', 8)]
        tissue = str(row[col.get('tissue', 4)]).strip() if row[col.get('tissue', 4)] else ""
        species = str(row[col.get('species', 0)]).strip() if col.get('species') is not None and row[col['species']] else ""
        if info == "None": info = ""
        if ref and el and eff is not None:
            try:
                gt[ref].append({
                    'element': el, 'effect': float(eff), 'info': info,
                    'eco2': float(eco2) if eco2 else None,
                    'tissue': tissue, 'species': species,
                })
            except (ValueError, TypeError):
                pass
    wb.close()
    return dict(gt)


def build_alignment_prompt(paper_id, gt_rows, obs_list, alignment_issue):
    """
    Build an LLM prompt to align extracted observations to GT rows.
    """
    # Format GT rows as a table
    gt_table = "GROUND TRUTH ROWS (what the meta-analyst selected):\n"
    gt_table += f"{'#':<4} {'Element':<6} {'Effect%':<10} {'Tissue':<12} {'Species':<25} {'Additional Info'}\n"
    gt_table += "-" * 90 + "\n"
    for i, gr in enumerate(gt_rows):
        gt_table += (f"{i:<4} {gr['element']:<6} {gr['effect']*100:>8.1f}%  "
                     f"{gr['tissue']:<12} {gr['species']:<25} {gr['info']}\n")

    # Format extracted observations
    obs_table = "\nEXTRACTED OBSERVATIONS (what the pipeline found):\n"
    obs_table += f"{'#':<4} {'Element':<6} {'T_mean':<10} {'C_mean':<10} {'Effect%':<10} {'Source':<15} {'Description'}\n"
    obs_table += "-" * 100 + "\n"
    for i, o in enumerate(obs_list):
        el = normalize_element(o.get('element', ''))
        t_mean = o.get('treatment_mean', '')
        c_mean = o.get('control_mean', '')
        if c_mean and c_mean != 0 and t_mean is not None:
            try:
                eff = (float(t_mean) - float(c_mean)) / float(c_mean) * 100
                eff_str = f"{eff:>8.1f}%"
            except:
                eff_str = "N/A"
        else:
            eff_str = "N/A"
        source = o.get('data_source', '')[:14]
        desc = str(o.get('treatment_description', ''))[:60]
        mods = o.get('moderators', {})
        if isinstance(mods, str):
            try: mods = json.loads(mods)
            except: mods = {}
        mod_str = ", ".join(f"{k}={v}" for k, v in (mods or {}).items() if v)[:60]
        obs_table += (f"{i:<4} {el:<6} {str(t_mean):<10} {str(c_mean):<10} "
                      f"{eff_str:<10} {source:<15} {desc}\n")
        if mod_str:
            obs_table += f"     Moderators: {mod_str}\n"

    prompt = f"""You are aligning extracted data to ground truth for validation of a meta-analysis extraction pipeline.

Paper: {paper_id}
Known alignment issue: {alignment_issue}

{gt_table}

{obs_table}

TASK: For each ground truth row (0 to {len(gt_rows)-1}), find the BEST matching extracted observation.
Match by: element, tissue type, sampling date, factorial condition, species, CO2 level.

The ground truth "Additional Info" column contains critical metadata like tissue type, sampling date (DOY),
cultivar, factorial level (+K/-K), ozone condition, soil type, etc. Use this to match precisely.

If NO extracted observation matches a GT row's conditions, output "NO_MATCH".
If the best match has a DIFFERENT tissue type, date, or factorial condition than what GT specifies,
output "MISALIGNED" with explanation.

Output as JSON array:
[
  {{"gt_row": 0, "best_obs": 5, "confidence": "high", "reason": "Same element Fe, same tissue grain, same year 2008"}},
  {{"gt_row": 1, "best_obs": null, "confidence": "none", "reason": "NO_MATCH - no Cu observations extracted"}},
  {{"gt_row": 2, "best_obs": 3, "confidence": "low", "reason": "MISALIGNED - GT wants September data, obs is from June"}}
]

Output ONLY the JSON array, nothing else.
"""
    return prompt


def run_alignment_llm(prompt):
    """Call Claude to align observations."""
    import openai
    client = openai.OpenAI(
        api_key=os.environ.get("ANTHROPIC_API_KEY"),
        base_url="https://api.anthropic.com/v1/"
    )

    response = client.chat.completions.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content


def run_alignment_gemini(prompt):
    """Call Gemini to align observations."""
    from google import genai
    client = genai.Client(api_key=os.environ.get("GOOGLE_API_KEY"))
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
    )
    return response.text


def compute_effect(obs):
    """Compute (E-A)/A."""
    ctrl = obs.get('control_mean')
    treat = obs.get('treatment_mean')
    if ctrl and ctrl != 0 and treat is not None:
        return (float(treat) - float(ctrl)) / float(ctrl)
    return None


def main():
    print("=" * 70)
    print("POST-EXTRACTION ALIGNMENT PASS")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 70)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    gt = load_gt()
    print(f"Loaded GT: {sum(len(v) for v in gt.values())} observations")

    # Process each alignment-problem paper
    all_alignments = {}
    total_improved = 0
    total_obs = 0

    for paper_id, issue in ALIGNMENT_PAPERS.items():
        loladze_ref = PAPER_TO_LOLADZE_REF.get(paper_id)
        if not loladze_ref or loladze_ref not in gt:
            print(f"\n  SKIP {paper_id}: no GT match for {loladze_ref}")
            continue

        # Load extracted observations
        consensus_file = RESULTS_DIR / f"{paper_id}_consensus.json"
        if not consensus_file.exists():
            print(f"\n  SKIP {paper_id}: no consensus file")
            continue

        with open(consensus_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        obs_list = data.get('consensus_observations', [])
        gt_rows = gt[loladze_ref]

        print(f"\n{'='*60}")
        print(f"Paper: {paper_id} ({loladze_ref})")
        print(f"Issue: {issue}")
        print(f"GT rows: {len(gt_rows)}, Extracted obs: {len(obs_list)}")

        # Build and run alignment prompt
        prompt = build_alignment_prompt(paper_id, gt_rows, obs_list, issue)

        print("  Calling Gemini for alignment...")
        try:
            response = run_alignment_gemini(prompt)
        except Exception as e:
            print(f"  ERROR: {e}")
            continue

        # Parse response
        try:
            # Extract JSON from response
            json_match = re.search(r'\[.*\]', response, re.DOTALL)
            if json_match:
                alignments = json.loads(json_match.group())
            else:
                print(f"  ERROR: No JSON found in response")
                print(f"  Response: {response[:200]}...")
                continue
        except json.JSONDecodeError as e:
            print(f"  ERROR parsing JSON: {e}")
            print(f"  Response: {response[:200]}...")
            continue

        # Analyze alignment results
        matched = 0
        misaligned = 0
        no_match = 0
        improved_matches = []

        for a in alignments:
            gt_idx = a.get('gt_row')
            obs_idx = a.get('best_obs')
            confidence = a.get('confidence', '')
            reason = a.get('reason', '')

            if obs_idx is None or 'NO_MATCH' in reason.upper():
                no_match += 1
            elif 'MISALIGNED' in reason.upper():
                misaligned += 1
            else:
                matched += 1
                # Compute error for this alignment
                if gt_idx is not None and gt_idx < len(gt_rows) and obs_idx is not None and obs_idx < len(obs_list):
                    gt_eff = gt_rows[gt_idx]['effect']
                    our_eff = compute_effect(obs_list[obs_idx])
                    if our_eff is not None:
                        err = abs(our_eff - gt_eff)
                        improved_matches.append({
                            'gt_row': gt_idx,
                            'obs_idx': obs_idx,
                            'gt_effect': gt_eff,
                            'our_effect': our_eff,
                            'err': err,
                            'element': gt_rows[gt_idx]['element'],
                            'confidence': confidence,
                            'reason': reason,
                        })

        print(f"  Results: {matched} matched, {misaligned} misaligned, {no_match} no-match")

        if improved_matches:
            mae = sum(m['err'] for m in improved_matches) / len(improved_matches) * 100
            print(f"  New MAE: {mae:.2f}% (from {len(improved_matches)} aligned matches)")
            total_obs += len(improved_matches)

        all_alignments[paper_id] = {
            'loladze_ref': loladze_ref,
            'issue': issue,
            'n_gt_rows': len(gt_rows),
            'n_extracted': len(obs_list),
            'matched': matched,
            'misaligned': misaligned,
            'no_match': no_match,
            'improved_matches': improved_matches,
            'raw_response': response,
        }

    # Save results
    # Strip raw responses for the summary file
    summary = {}
    for pid, data in all_alignments.items():
        d = dict(data)
        del d['raw_response']
        summary[pid] = d

    with open(OUTPUT_DIR / "alignment_results.json", 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, default=str)

    # Also save full results with LLM responses
    with open(OUTPUT_DIR / "alignment_full.json", 'w', encoding='utf-8') as f:
        json.dump(all_alignments, f, indent=2, default=str)

    # Compute overall improvement
    print(f"\n{'='*70}")
    print("ALIGNMENT PASS SUMMARY")
    print(f"{'='*70}")
    print(f"Papers processed: {len(all_alignments)}")
    print(f"Total aligned observations: {total_obs}")

    all_improved = []
    for pid, data in all_alignments.items():
        for m in data.get('improved_matches', []):
            all_improved.append(m)

    if all_improved:
        new_mae = sum(m['err'] for m in all_improved) / len(all_improved) * 100
        print(f"Aligned MAE (problem papers only): {new_mae:.2f}%")

    print(f"\nSaved: {OUTPUT_DIR / 'alignment_results.json'}")
    print(f"Saved: {OUTPUT_DIR / 'alignment_full.json'}")
    print("\nNext step: integrate these alignments into validate_full_46.py")


if __name__ == "__main__":
    main()
